import logging
#!/usr/bin/env python3
"""
خدمة التقارير - Reports Service
تدير إنشاء وتصدير التقارير المختلفة للنظام
"""

import json
import os
from dataclasses import asdict, dataclass
from datetime import datetime
from enum import Enum
from typing import Any, Dict, List, Optional

# استيراد المكتبات المطلوبة للتصدير
try:
    import pandas as pd
    from openpyxl import Workbook  # pyright: ignore[reportMissingModuleSource]
    from openpyxl.styles import (  # pyright: ignore[reportMissingModuleSource]
        Alignment,
        Font,
        PatternFill,
    )
    from openpyxl.utils.dataframe import (
        dataframe_to_rows,
    )  # pyright: ignore[reportMissingModuleSource]

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

import os  # noqa: F811
import warnings

from jinja2 import Template

# قمع تحذيرات WeasyPrint قبل الاستيراد
warnings.filterwarnings("ignore", category=UserWarning, module="weasyprint")
warnings.filterwarnings("ignore", message=".*WeasyPrint.*", category=UserWarning)

try:
    from weasyprint import CSS, HTML  # type: ignore

    PDF_AVAILABLE = True
except (ImportError, OSError, RuntimeError):
    # يشمل ImportError و OSError الناتج عن عدم توفر مكتبات نظام مثل gobject/pango
    # RuntimeError قد يحدث عند عدم توفر مكتبات النظام المطلوبة
    PDF_AVAILABLE = False
    HTML = CSS = None  # تعيين قيم افتراضية آمنة

try:
    from pptx import Presentation  # pyright: ignore[reportMissingModuleSource]

    PPTX_AVAILABLE = True
except ImportError:
    PPTX_AVAILABLE = False
    Presentation = None  # تعيين قيمة افتراضية آمنة

try:
    # محاولة الاستيراد النسبي أولاً (عند الاستيراد من داخل الحزمة)
    from ..core.database_manager import DatabaseManager
    from ..models.customer import CustomerManager
    from ..models.product import ProductManager
    from ..models.purchase import PurchaseManager
    from ..models.report import ExportFormat, ReportData, ReportFilter, ReportType
    from ..models.sale import SaleManager
    from ..models.supplier import SupplierManager
    from ..utils.db_helpers import get_value
    from ..utils.logger import DatabaseLogger, setup_logger
except ImportError:
    # Fallback للاستيراد المطلق (عند الاستيراد المباشر)
    try:
        from src.core.database_manager import DatabaseManager
        from src.models.customer import CustomerManager
        from src.models.product import ProductManager
        from src.models.purchase import PurchaseManager
        from src.models.report import ExportFormat, ReportData, ReportFilter, ReportType
        from src.models.sale import SaleManager
        from src.models.supplier import SupplierManager
        from src.utils.db_helpers import get_value
        from src.utils.logger import DatabaseLogger, setup_logger
    except ImportError:
        # Fallback: تعريف محلي إذا لم يكن models.report متوفراً
        class ReportType(Enum):
            """أنواع التقارير المتاحة"""

            SALES_SUMMARY = "sales_summary"
            INVENTORY_STATUS = "inventory_status"
            FINANCIAL_SUMMARY = "financial_summary"
            PAYMENT_SUMMARY = "payment_summary"
            RECEIVABLES_AGING = "receivables_aging"
            PAYABLES_AGING = "payables_aging"
            CASH_FLOW = "cash_flow"
            PAYMENT_ANALYSIS = "payment_analysis"
            PAYMENT_METHODS_ANALYSIS = "payment_methods_analysis"
            # أنواع إضافية مستخدمة في reports_window
            CUSTOMER_ANALYSIS = "customer_analysis"
            SUPPLIER_ANALYSIS = "supplier_analysis"
            PRODUCT_PERFORMANCE = "product_performance"
            PROFIT_LOSS = "profit_loss"
            STOCK_MOVEMENT = "stock_movement"
            # Multi-Warehouse Reports
            WAREHOUSE_INVENTORY = "warehouse_inventory"
            WAREHOUSE_TRANSFERS = "warehouse_transfers"
            WAREHOUSE_LOW_STOCK = "warehouse_low_stock"
            WAREHOUSE_SUMMARY = "warehouse_summary"

    @dataclass
    class ReportFilter:  # noqa: F811
        """فلاتر التقارير"""

        start_date: Optional[datetime] = None
        end_date: Optional[datetime] = None
        customer_id: Optional[int] = None
        product_id: Optional[int] = None
        category_id: Optional[int] = None
        supplier_id: Optional[int] = None
        user_id: Optional[int] = None
        entity_id: Optional[int] = None
        payment_method: Optional[str] = None
        payment_type: Optional[str] = None
        payment_status: Optional[str] = None
        account_type: Optional[str] = None
        group_by: Optional[str] = None
        min_amount: Optional[float] = None
        max_amount: Optional[float] = None
        aging_periods: Optional[List[int]] = None
        include_zero_balances: bool = True
        warehouse_id: Optional[int] = None  # Multi-Warehouse Support

    @dataclass
    class ReportData:  # noqa: F811
        """بيانات التقرير"""

        title: str
        subtitle: str
        generated_at: datetime
        filters: ReportFilter
        data: List[Dict[str, Any]]
        summary: Dict[str, Any]
        charts_data: Optional[Dict[str, Any]] = None

    class ExportFormat(Enum):  # noqa: F811
        """صيغ التصدير"""

        PDF = "pdf"
        EXCEL = "excel"
        CSV = "csv"
        JSON = "json"
        PPTX = "pptx"


class ReportExporter:
    """خدمة التقارير"""

    def __init__(self, db_manager: DatabaseManager):
        self.db = db_manager
        self.db_manager = db_manager
        self.logger = setup_logger(__name__)
        self.db_logger = DatabaseLogger(db_manager)

        # إنشاء مديري النماذج
        self.product_manager = ProductManager(db_manager)
        self.sale_manager = SaleManager(db_manager)
        self.purchase_manager = PurchaseManager(db_manager)
        self.customer_manager = CustomerManager(db_manager)
        self.supplier_manager = SupplierManager(db_manager)

        # Multi-Currency Support
        try:
            from ..models.currency import CurrencyManager
            from ..services.exchange_rate_service import ExchangeRateService

            self.currency_manager = CurrencyManager(db_manager)
            self.exchange_rate_service = ExchangeRateService(db_manager, self.logger)
        except ImportError:
            self.currency_manager = None
            self.exchange_rate_service = None

        # إعداد مجلد التقارير
        self.reports_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
            "data",
            "exports",
        )
        os.makedirs(self.reports_dir, exist_ok=True)

    def generate_report(self, report_type: ReportType, filters: ReportFilter) -> ReportData:
        """توليد تقرير بناءً على نوع التقرير المحدد"""
        try:
            if report_type == ReportType.SALES_SUMMARY:
                return self.generate_sales_summary_report(filters)
            elif report_type == ReportType.INVENTORY_STATUS:
                return self.generate_inventory_status_report(filters)
            elif report_type == ReportType.FINANCIAL_SUMMARY:
                return self.generate_financial_summary_report(filters)
            elif report_type == ReportType.PAYMENT_SUMMARY:
                return self.generate_payment_summary_report(filters)
            elif report_type == ReportType.RECEIVABLES_AGING:
                return self.generate_receivables_aging_report(filters)
            elif report_type == ReportType.PAYABLES_AGING:
                return self.generate_payables_aging_report(filters)
            elif report_type == ReportType.CASH_FLOW:
                return self.generate_cash_flow_report(filters)
            elif report_type == ReportType.PAYMENT_ANALYSIS:
                return self.generate_payment_analysis_report(filters)
            elif report_type == ReportType.PAYMENT_METHODS_ANALYSIS:
                return self.generate_payment_methods_analysis_report(filters)
            elif report_type == ReportType.CUSTOMER_ANALYSIS:
                return self.generate_customer_analysis_report(filters)
            elif report_type == ReportType.SUPPLIER_ANALYSIS:
                return self.generate_supplier_analysis_report(filters)
            elif report_type == ReportType.PRODUCT_PERFORMANCE:
                return self.generate_product_performance_report(filters)
            elif report_type == ReportType.PROFIT_LOSS:
                # استخدام التقرير المالي كبديل مؤقت
                return self.generate_financial_summary_report(filters)
            elif report_type == ReportType.STOCK_MOVEMENT:
                return self.generate_stock_movement_report(filters)
            elif report_type == ReportType.WAREHOUSE_INVENTORY:
                return self.generate_warehouse_inventory_report(filters)
            elif report_type == ReportType.WAREHOUSE_TRANSFERS:
                return self.generate_warehouse_transfers_report(filters)
            elif report_type == ReportType.WAREHOUSE_LOW_STOCK:
                return self.generate_warehouse_low_stock_report(filters)
            elif report_type == ReportType.WAREHOUSE_SUMMARY:
                return self.generate_warehouse_summary_report(filters)
            else:
                raise ValueError(f"نوع التقرير غير مدعوم: {report_type}")
        except Exception as e:
            self.logger.error(f"خطأ في توليد التقرير {report_type}: {e}")
            raise

    def generate_sales_summary_report(self, filters: ReportFilter) -> ReportData:
        """إنشاء تقرير ملخص المبيعات"""
        try:
            # بناء استعلام المبيعات (مع دعم Multi-Currency)
            query = """
                SELECT
                    s.id,
                    s.invoice_number,
                    s.sale_date,
                    s.total_amount,
                    s.discount_amount,
                    s.tax_amount,
                    s.final_amount,
                    s.payment_method,
                    s.status,
                    s.currency_id,
                    s.exchange_rate,
                    s.base_amount,
                    s.converted_amount,
                    c.name as customer_name,
                    u.username as user_name,
                    curr.code as currency_code,
                    curr.name as currency_name,
                    curr.symbol as currency_symbol
                FROM sales s
                LEFT JOIN customers c ON s.customer_id = c.id
                LEFT JOIN users u ON s.user_id = u.id
                LEFT JOIN currencies curr ON s.currency_id = curr.id
                WHERE 1=1
            """

            params = []

            # تطبيق الفلاتر
            if filters.start_date:
                query += " AND s.sale_date >= ?"
                params.append(filters.start_date.strftime("%Y-%m-%d"))

            if filters.end_date:
                query += " AND s.sale_date <= ?"
                params.append(filters.end_date.strftime("%Y-%m-%d"))

            if filters.customer_id:
                query += " AND s.customer_id = ?"
                params.append(filters.customer_id)

            if filters.user_id:
                query += " AND s.user_id = ?"
                params.append(filters.user_id)

            if filters.min_amount:
                query += " AND s.final_amount >= ?"
                params.append(filters.min_amount)

            if filters.max_amount:
                query += " AND s.final_amount <= ?"
                params.append(filters.max_amount)

            query += " ORDER BY s.sale_date DESC"

            # تنفيذ الاستعلام - استخدام execute_query للحصول على dictionaries
            rows = self.db.execute_query(query, tuple(params))

            # تحويل النتائج إلى قواميس مع تحويل الحالة إلى العربية
            def translate_status(status):
                """تحويل حالة الفاتورة من الإنجليزية إلى العربية"""
                if not status:
                    return "-"
                status_lower = str(status).lower().strip()
                status_map = {
                    "draft": "مسودة",
                    "pending": "مؤكدة",
                    "confirmed": "مؤكدة",
                    "paid": "مدفوعة",
                    "partially_paid": "مدفوعة جزئياً",
                    "cancelled": "ملغية",
                    "returned": "مرتجعة",
                    "مسودة": "مسودة",
                    "مؤكدة": "مؤكدة",
                    "مدفوعة": "مدفوعة",
                    "مدفوعة جزئياً": "مدفوعة جزئياً",
                    "ملغية": "ملغية",
                    "مرتجعة": "مرتجعة",
                }
                return status_map.get(status_lower, status)

            def translate_payment_method(method):
                """تحويل طريقة الدفع من الإنجليزية إلى العربية"""
                if not method:
                    return "-"
                method_lower = str(method).lower().strip()
                method_map = {
                    "cash": "نقدي",
                    "card": "بطاقة بنكية",
                    "bank_transfer": "تحويل بنكي",
                    "credit": "آجل (ذمم)",
                    "نقدي": "نقدي",
                    "بطاقة": "بطاقة بنكية",
                    "بطاقة بنكية": "بطاقة بنكية",
                    "تحويل": "تحويل بنكي",
                    "تحويل بنكي": "تحويل بنكي",
                    "آجل": "آجل (ذمم)",
                    "آجل (ذمم)": "آجل (ذمم)",
                }
                return method_map.get(method_lower, method)

            # تحويل الصفوف إلى قواميس مع تحويل الحالة وطريقة الدفع
            sales_data = []
            for row in rows:
                currency_id = row.get("currency_id")
                currency_code = row.get("currency_code") or "DZD"
                currency_name = row.get("currency_name") or "الدينار الجزائري"
                currency_symbol = row.get("currency_symbol") or "د.ج"
                base_amount = row.get("base_amount")
                converted_amount = row.get("converted_amount")
                exchange_rate = row.get("exchange_rate") or 1.0

                # استخدام base_amount إذا كان متوفراً، وإلا استخدام final_amount
                display_amount = float(base_amount) if base_amount else float(row.get("final_amount") or 0)

                # row هو dictionary من execute_query
                sales_data.append(
                    {
                        "id": row.get("id"),
                        "invoice_number": row.get("invoice_number", ""),
                        "sale_date": row.get("sale_date"),
                        "total_amount": float(row.get("total_amount") or 0),
                        "discount_amount": float(row.get("discount_amount") or 0),
                        "tax_amount": float(row.get("tax_amount") or 0),
                        "final_amount": float(row.get("final_amount") or 0),
                        # Multi-Currency Support
                        "currency_id": currency_id,
                        "currency_code": currency_code,
                        "currency_name": currency_name,
                        "currency_symbol": currency_symbol,
                        "base_amount": (float(base_amount) if base_amount else display_amount),
                        "converted_amount": (float(converted_amount) if converted_amount else display_amount),
                        "exchange_rate": float(exchange_rate),
                        "display_amount": display_amount,  # المبلغ للعرض (بالعملة الأساسية)
                        "payment_method": translate_payment_method(row.get("payment_method")),
                        "status": translate_status(row.get("status")),
                        "customer_name": row.get("customer_name") or "",
                        "user_name": row.get("user_name") or "",
                    }
                )

            # حساب الملخص (باستخدام المبالغ بالعملة الأساسية)
            total_sales = len(sales_data)
            total_amount = sum(row.get("base_amount", row["final_amount"]) for row in sales_data)
            total_discount = sum(row["discount_amount"] or 0 for row in sales_data)
            total_tax = sum(row["tax_amount"] or 0 for row in sales_data)

            # تجميع حسب العملة
            currencies_summary = {}
            for row in sales_data:
                currency_code = row.get("currency_code", "DZD")
                if currency_code not in currencies_summary:
                    currencies_summary[currency_code] = {
                        "count": 0,
                        "amount": 0,
                        "base_amount": 0,
                    }
                currencies_summary[currency_code]["count"] += 1
                currencies_summary[currency_code]["amount"] += row.get("converted_amount", row["final_amount"])
                currencies_summary[currency_code]["base_amount"] += row.get("base_amount", row["final_amount"])

            # تجميع البيانات حسب طريقة الدفع
            payment_methods = {}
            for row in sales_data:
                method = row["payment_method"]
                if method not in payment_methods:
                    payment_methods[method] = {"count": 0, "amount": 0}
                payment_methods[method]["count"] += 1
                payment_methods[method]["amount"] += row["final_amount"]

            summary = {
                "total_sales": total_sales,
                "total_amount": total_amount,  # بالعملة الأساسية
                "total_discount": total_discount,
                "total_tax": total_tax,
                "average_sale": total_amount / total_sales if total_sales > 0 else 0,
                "payment_methods": payment_methods,
                # Multi-Currency Support
                "currencies_summary": currencies_summary,
            }

            # بيانات الرسوم البيانية
            charts_data = {
                "daily_sales": self._get_daily_sales_chart_data(filters),
                "payment_methods": payment_methods,
                "top_customers": self._get_top_customers_chart_data(filters),
            }

            return ReportData(
                title="تقرير ملخص المبيعات",
                subtitle=f"من {filters.start_date.strftime('%Y-%m-%d') if filters.start_date else 'البداية'} إلى {filters.end_date.strftime('%Y-%m-%d') if filters.end_date else 'النهاية'}",  # noqa: E501
                generated_at=datetime.now(),
                filters=filters,
                data=sales_data,
                summary=summary,
                charts_data=charts_data,
            )

        except Exception as e:
            self.logger.error(f"خطأ في إنشاء تقرير ملخص المبيعات: {e}")
            raise

    def generate_inventory_status_report(self, filters: ReportFilter) -> ReportData:
        """إنشاء تقرير حالة المخزون"""
        try:
            # التحقق من وجود الأعمدة في جدول products
            try:
                columns_info = self.db.fetch_all("PRAGMA table_info(products)")
                # PRAGMA table_info returns tuples: (cid, name, type, notnull, default_value, pk)
                available_columns = {row[1] for row in columns_info} if columns_info else set()
            except Exception:
                # في حالة الخطأ، افترض أن الأعمدة الأساسية موجودة فقط
                available_columns = {
                    "id",
                    "name",
                    "barcode",
                    "current_stock",
                    "min_stock",
                    "cost_price",
                    "selling_price",
                    "category_id",
                    "is_active",
                }

            # تحديد الأعمدة المتاحة
            has_supplier_id = "supplier_id" in available_columns
            has_max_stock = "max_stock" in available_columns

            # بناء الاستعلام بناءً على الأعمدة المتاحة
            supplier_join = "LEFT JOIN suppliers s ON p.supplier_id = s.id" if has_supplier_id else ""  # noqa: F841
            supplier_select = "s.name as supplier_name," if has_supplier_id else "NULL as supplier_name,"  # noqa: F841
            "COALESCE(p.max_stock, 0)" if has_max_stock else "0"

            query = """
                SELECT
                    p.id,
                    p.name,
                    p.barcode,
                    p.current_stock,
                    COALESCE(p.min_stock, 0) as min_stock_level,
                    {max_stock_col} as max_stock_level,
                    COALESCE(p.cost_price, 0) as cost_price,
                    COALESCE(p.selling_price, 0) as selling_price,
                    p.current_stock * COALESCE(p.cost_price, 0) as stock_value,
                    c.name as category_name,
                    {supplier_select}
                    CASE
                        WHEN p.current_stock <= 0 THEN 'نفد المخزون'
                        WHEN COALESCE(p.min_stock, 0) > 0 AND p.current_stock <= COALESCE(p.min_stock, 0) THEN 'مخزون منخفض'  # noqa: E501
                        WHEN {max_stock_col} > 0 AND p.current_stock >= {max_stock_col} THEN 'مخزون مرتفع'
                        ELSE 'طبيعي'
                    END as stock_status
                FROM products p
                LEFT JOIN categories c ON p.category_id = c.id
                {supplier_join}
                WHERE p.is_active = 1
            """

            params = []

            if filters.category_id:
                query += " AND p.category_id = ?"
                params.append(filters.category_id)

            if filters.supplier_id and has_supplier_id:
                query += " AND p.supplier_id = ?"
                params.append(filters.supplier_id)

            query += " ORDER BY p.name"

            products_data = self.db.execute_query(query, params)

            # حساب الملخص
            total_products = len(products_data)
            total_stock_value = sum(float(row.get("stock_value", 0) or 0) for row in products_data)
            out_of_stock = len([p for p in products_data if (p.get("current_stock") or 0) <= 0])
            low_stock = len(
                [
                    p
                    for p in products_data
                    if (p.get("current_stock") or 0) > 0
                    and (p.get("min_stock_level") or 0) > 0
                    and (p.get("current_stock") or 0) <= (p.get("min_stock_level") or 0)
                ]
            )
            high_stock = len(
                [
                    p
                    for p in products_data
                    if (p.get("max_stock_level") or 0) > 0
                    and (p.get("current_stock") or 0) >= (p.get("max_stock_level") or 0)
                ]
            )

            # تجميع حسب الفئات
            categories_summary = {}
            for row in products_data:
                cat = row["category_name"] or "بدون فئة"
                if cat not in categories_summary:
                    categories_summary[cat] = {"count": 0, "value": 0}
                categories_summary[cat]["count"] += 1
                categories_summary[cat]["value"] += row["stock_value"]

            summary = {
                "total_products": total_products,
                "total_stock_value": total_stock_value,
                "out_of_stock": out_of_stock,
                "low_stock": low_stock,
                "high_stock": high_stock,
                "normal_stock": total_products - out_of_stock - low_stock - high_stock,
                "categories_summary": categories_summary,
            }

            charts_data = {
                "stock_status": {
                    "نفد المخزون": out_of_stock,
                    "مخزون منخفض": low_stock,
                    "مخزون مرتفع": high_stock,
                    "طبيعي": summary["normal_stock"],
                },
                "categories_value": categories_summary,
            }

            return ReportData(
                title="تقرير حالة المخزون",
                subtitle="حالة المخزون الحالية لجميع المنتجات",
                generated_at=datetime.now(),
                filters=filters,
                data=products_data,
                summary=summary,
                charts_data=charts_data,
            )

        except Exception as e:
            self.logger.error(f"خطأ في إنشاء تقرير حالة المخزون: {e}")
            raise

    def generate_financial_summary_report(self, filters: ReportFilter) -> ReportData:
        """إنشاء تقرير الملخص المالي"""
        try:
            # المبيعات (باستخدام base_amount للعملة الأساسية)
            sales_query = """
                SELECT
                    SUM(COALESCE(base_amount, final_amount)) as total_sales,
                    COUNT(*) as sales_count
                FROM sales
                WHERE status != 'cancelled'
            """

            # المشتريات (باستخدام base_amount للعملة الأساسية)
            purchases_query = """
                SELECT
                    SUM(COALESCE(base_amount, total_amount)) as total_purchases,
                    COUNT(*) as purchases_count
                FROM purchases
                WHERE status != 'cancelled'
            """

            # الربح الفعلي من بنود المبيعات
            profit_query = """
                SELECT SUM(si.profit)
                FROM sale_items si
                JOIN sales s ON si.sale_id = s.id
                WHERE s.status != 'cancelled'
            """

            date_params = []

            if filters.start_date:
                sales_query += " AND sale_date >= ?"
                purchases_query += " AND purchase_date >= ?"
                profit_query += " AND s.sale_date >= ?"
                date_params.append(filters.start_date.strftime("%Y-%m-%d"))

            if filters.end_date:
                sales_query += " AND sale_date <= ?"
                purchases_query += " AND purchase_date <= ?"
                profit_query += " AND s.sale_date <= ?"
                date_params.append(filters.end_date.strftime("%Y-%m-%d"))

            # تنفيذ الاستعلامات
            sales_result = self.db.fetch_one(sales_query, tuple(date_params))
            purchases_result = self.db.fetch_one(purchases_query, tuple(date_params))
            gross_profit = self.db.execute_scalar(profit_query, tuple(date_params)) or 0

            sales_data = {
                "total_sales": sales_result[0] or 0,
                "sales_count": sales_result[1] or 0,
            }
            purchases_data = {
                "total_purchases": purchases_result[0] or 0,
                "purchases_count": purchases_result[1] or 0,
            }

            total_sales = sales_data["total_sales"]

            # حساب قيمة المخزون الحالي
            inventory_value = (
                self.db.execute_scalar("SELECT SUM(current_stock * cost_price) FROM products WHERE is_active = 1") or 0
            )

            summary = {
                "total_sales": total_sales,
                "total_purchases": purchases_data["total_purchases"],
                "gross_profit": gross_profit,
                "profit_margin": ((gross_profit / total_sales * 100) if total_sales > 0 else 0),
                "inventory_value": inventory_value,
                "sales_count": sales_data["sales_count"],
                "purchases_count": purchases_data["purchases_count"],
            }

            detailed_data = [
                {
                    "metric": "إجمالي المبيعات",
                    "value": summary["total_sales"],
                    "type": "income",
                },
                {
                    "metric": "إجمالي المشتريات (للفترة)",
                    "value": summary["total_purchases"],
                    "type": "expense",
                },
                {
                    "metric": "الربح الإجمالي الفعلي",
                    "value": summary["gross_profit"],
                    "type": "profit",
                },
                {
                    "metric": "قيمة المخزون الحالية",
                    "value": summary["inventory_value"],
                    "type": "asset",
                },
            ]

            charts_data = {
                "profit_breakdown": {"المبيعات": total_sales, "الربح": gross_profit},
                "monthly_trend": self._get_monthly_financial_trend(filters),
            }

            return ReportData(
                title="تقرير الملخص المالي",
                subtitle=f"من {filters.start_date.strftime('%Y-%m-%d') if filters.start_date else 'البداية'} إلى {filters.end_date.strftime('%Y-%m-%d') if filters.end_date else 'النهاية'}",  # noqa: E501
                generated_at=datetime.now(),
                filters=filters,
                data=detailed_data,
                summary=summary,
                charts_data=charts_data,
            )

        except Exception as e:
            self.logger.error(f"خطأ في إنشاء تقرير الملخص المالي: {e}")
            raise

    def export_report(
        self,
        report_data: ReportData,
        format_type: ExportFormat,
        filename: Optional[str] = None,
    ) -> str:
        """تصدير التقرير بالتنسيق المحدد"""
        try:
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"report_{timestamp}.{format_type.value}"

            filepath = os.path.join(self.reports_dir, filename)

            if format_type == ExportFormat.JSON:
                return self._export_to_json(report_data, filepath)
            elif format_type == ExportFormat.CSV:
                return self._export_to_csv(report_data, filepath)
            elif format_type == ExportFormat.EXCEL and EXCEL_AVAILABLE:
                return self._export_to_excel(report_data, filepath)
            elif format_type == ExportFormat.PDF and PDF_AVAILABLE:
                return self._export_to_pdf(report_data, filepath)
            elif format_type == ExportFormat.PPTX and PPTX_AVAILABLE:
                return self._export_to_pptx(report_data, filepath)
            else:
                raise ValueError(f"تنسيق التصدير غير مدعوم أو المكتبات المطلوبة غير متوفرة: {format_type.value}")

        except Exception as e:
            self.logger.error(f"خطأ في تصدير التقرير: {e}")
            raise

    def export_to_pdf(self, data: List[Dict[str, Any]], title: str, filepath: str) -> str:
        """
        تصدير بيانات إلى PDF (wrapper method للاستخدام البسيط)

        Args:
            data: قائمة من القواميس (البيانات)
            title: عنوان التقرير
            filepath: مسار الملف النهائي
        """
        # إنشاء ReportData object من البيانات البسيطة
        report_data = ReportData(
            title=title,
            subtitle="",
            generated_at=datetime.now(),
            filters=ReportFilter(),
            data=data,
            summary={},
        )
        return self._export_to_pdf(report_data, filepath)

    def export_to_excel(self, data: List[Dict[str, Any]], filepath: str) -> str:
        """
        تصدير بيانات إلى Excel (wrapper method للاستخدام البسيط)

        Args:
            data: قائمة من القواميس (البيانات)
            filepath: مسار الملف النهائي
        """
        # إنشاء ReportData object من البيانات البسيطة
        report_data = ReportData(
            title="تقرير",
            subtitle="",
            generated_at=datetime.now(),
            filters=ReportFilter(),
            data=data,
            summary={},
        )
        return self._export_to_excel(report_data, filepath)

    def export_to_csv(self, data: List[Dict[str, Any]], filepath: str) -> str:
        """
        تصدير بيانات إلى CSV (wrapper method للاستخدام البسيط)

        Args:
            data: قائمة من القواميس (البيانات)
            filepath: مسار الملف النهائي
        """
        # إنشاء ReportData object من البيانات البسيطة
        report_data = ReportData(
            title="تقرير",
            subtitle="",
            generated_at=datetime.now(),
            filters=ReportFilter(),
            data=data,
            summary={},
        )
        return self._export_to_csv(report_data, filepath)

    def _export_to_json(self, report_data: ReportData, filepath: str) -> str:
        """تصدير إلى JSON"""
        data = asdict(report_data)
        # تحويل التواريخ والـ Enums إلى نصوص
        data["generated_at"] = report_data.generated_at.isoformat()
        if data["filters"].get("start_date"):
            data["filters"]["start_date"] = report_data.filters.start_date.isoformat()
        if data["filters"].get("end_date"):
            data["filters"]["end_date"] = report_data.filters.end_date.isoformat()

        # تحويل الـ Enums في الفلاتر (مثل ReportPeriod)
        from enum import Enum

        if "filters" in data:
            for key, value in data["filters"].items():
                if isinstance(value, Enum):
                    data["filters"][key] = value.value

        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        return filepath

    def _export_to_csv(self, report_data: ReportData, filepath: str) -> str:
        """تصدير إلى CSV"""
        if not report_data.data:
            raise ValueError("لا توجد بيانات للتصدير")

        df = pd.DataFrame(report_data.data)
        df.to_csv(filepath, index=False, encoding="utf-8-sig")

        return filepath

    def _export_to_excel(self, report_data: ReportData, filepath: str) -> str:
        """تصدير إلى Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "التقرير"

        # إعداد الأنماط
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # كتابة العنوان
        ws["A1"] = report_data.title
        ws["A1"].font = Font(bold=True, size=16)
        ws["A2"] = report_data.subtitle
        ws["A3"] = f"تاريخ الإنشاء: {report_data.generated_at.strftime('%Y-%m-%d %H:%M:%S')}"

        # كتابة البيانات
        if report_data.data:
            df = pd.DataFrame(report_data.data)

            # إضافة البيانات بدءاً من الصف 5
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            # تنسيق الرأس
            for cell in ws[5]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

        # إضافة ورقة الملخص
        summary_ws = wb.create_sheet("الملخص")
        row = 1
        for key, value in report_data.summary.items():
            if isinstance(value, dict):
                summary_ws[f"A{row}"] = str(key)
                summary_ws[f"A{row}"].font = Font(bold=True)
                row += 1
                for sub_key, sub_value in value.items():
                    summary_ws[f"A{row}"] = f"  {sub_key}"
                    summary_ws[f"B{row}"] = str(sub_value)
                    row += 1
            else:
                summary_ws[f"A{row}"] = str(key)
                summary_ws[f"B{row}"] = str(value)
                row += 1

        wb.save(filepath)
        return filepath

    def _export_to_pdf(self, report_data: ReportData, filepath: str) -> str:
        """تصدير إلى PDF"""
        if not PDF_AVAILABLE or HTML is None:
            raise RuntimeError("مكتبات WeasyPrint غير متوفرة في هذا البيئة، لا يمكن التصدير إلى PDF")
        # قالب HTML للتقرير
        html_template = """
        <!DOCTYPE html>
        <html dir="rtl" lang="ar">
        <head>
            <meta charset="UTF-8">
            <title>{{ title }}</title>
            <style>
                body { font-family: 'Arial', sans-serif; direction: rtl; }
                .header { text-align: center; margin-bottom: 30px; }
                .title { font-size: 24px; font-weight: bold; color: #333; }
                .subtitle { font-size: 16px; color: #666; margin-top: 10px; }
                .summary { background: #f5f5f5; padding: 15px; margin: 20px 0; border-radius: 5px; }
                .data-table { width: 100%; border-collapse: collapse; margin-top: 20px; }
                .data-table th, .data-table td { border: 1px solid #ddd; padding: 8px; text-align: right; }
                .data-table th { background-color: #366092; color: white; }
                .footer { margin-top: 30px; text-align: center; font-size: 12px; color: #666; }
            </style>
        </head>
        <body>
            <div class="header">
                <div class="title">{{ title }}</div>
                <div class="subtitle">{{ subtitle }}</div>
                <div>تاريخ الإنشاء: {{ generated_at }}</div>
            </div>

            <div class="summary">
                <h3>ملخص التقرير</h3>
                {% for key, value in summary.items() %}
                    <p><strong>{{ key }}:</strong> {{ value }}</p>
                {% endfor %}
            </div>

            {% if data %}
            <table class="data-table">
                <thead>
                    <tr>
                        {% for key in data[0].keys() %}
                            <th>{{ key }}</th>
                        {% endfor %}
                    </tr>
                </thead>
                <tbody>
                    {% for row in data %}
                    <tr>
                        {% for value in row.values() %}
                            <td>{{ value }}</td>
                        {% endfor %}
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
            {% endif %}

            <div class="footer">
                <p>تم إنشاء هذا التقرير بواسطة نظام الإصدار المنطقي</p>
            </div>
        </body>
        </html>
        """

        template = Template(html_template)
        html_content = template.render(
            title=report_data.title,
            subtitle=report_data.subtitle,
            generated_at=report_data.generated_at.strftime("%Y-%m-%d %H:%M:%S"),
            summary=report_data.summary,
            data=report_data.data,
        )

        # تحويل إلى PDF
        HTML(string=html_content).write_pdf(filepath)
        return filepath

    def _export_to_pptx(self, report_data: ReportData, filepath: str) -> str:
        """تصدير إلى PPTX"""
        if not PPTX_AVAILABLE or Presentation is None:
            raise RuntimeError("مكتبات python-pptx غير متوفرة")

        try:
            from .pptx_report_builder import PptxReportBuilder, PptxReportTheme
        except ImportError:
            from src.services.pptx_report_builder import (
                PptxReportBuilder,
                PptxReportTheme,
            )

        builder = PptxReportBuilder(PptxReportTheme.CORPORATE)
        builder.add_title_slide(
            report_data.title,
            report_data.subtitle,
            "Logical Version ERP",
            report_data.generated_at.strftime("%Y-%m-%d"),
        )

        if report_data.summary:
            summary_items = [f"{k}: {v}" for k, v in report_data.summary.items()]
            builder.add_summary_slide(summary_items)

        builder.add_content_slide(
            report_data.title,
            {
                "layout": "list",
                "body": (report_data.data[:20] if len(report_data.data) > 20 else report_data.data),
            },
        )

        return builder.generate_pptx(filepath)

    def _get_daily_sales_chart_data(self, filters: ReportFilter) -> Dict[str, float]:
        """الحصول على بيانات المبيعات اليومية للرسم البياني"""
        query = """
            SELECT
                DATE(sale_date) as sale_day,
                SUM(final_amount) as daily_total
            FROM sales
            WHERE status != 'cancelled'
        """

        params = []
        if filters.start_date:
            query += " AND sale_date >= ?"
            params.append(filters.start_date.strftime("%Y-%m-%d"))

        if filters.end_date:
            query += " AND sale_date <= ?"
            params.append(filters.end_date.strftime("%Y-%m-%d"))

        query += " GROUP BY DATE(sale_date) ORDER BY sale_day"

        results = self.db.execute_query(query, params)
        return {row["sale_day"]: row["daily_total"] for row in results}

    def _get_top_customers_chart_data(self, filters: ReportFilter) -> Dict[str, float]:
        """الحصول على بيانات أفضل العملاء للرسم البياني"""
        query = """
            SELECT
                c.name as customer_name,
                SUM(s.final_amount) as total_purchases
            FROM sales s
            JOIN customers c ON s.customer_id = c.id
            WHERE s.status != 'cancelled'
        """

        params = []
        if filters.start_date:
            query += " AND s.sale_date >= ?"
            params.append(filters.start_date.strftime("%Y-%m-%d"))

        if filters.end_date:
            query += " AND s.sale_date <= ?"
            params.append(filters.end_date.strftime("%Y-%m-%d"))

        query += " GROUP BY c.id, c.name ORDER BY total_purchases DESC LIMIT 10"

        results = self.db.execute_query(query, params)
        return {row["customer_name"]: row["total_purchases"] for row in results}

    def _get_monthly_financial_trend(self, filters: ReportFilter) -> Dict[str, Dict[str, float]]:
        """الحصول على الاتجاه المالي الشهري"""
        date_params = []
        start_date_str = ""
        end_date_str = ""

        if filters.start_date:
            start_date_str = " AND s.sale_date >= ? "  # noqa: F841
            date_params.append(filters.start_date.strftime("%Y-%m-%d"))
        if filters.end_date:
            end_date_str = " AND s.sale_date <= ? "  # noqa: F841
            date_params.append(filters.end_date.strftime("%Y-%m-%d"))

        # الربح الشهري الفعلي
        profit_query = """
            SELECT
                strftime('%Y-%m', s.sale_date) as month,
                SUM(si.profit) as monthly_profit
            FROM sale_items si
            JOIN sales s ON si.sale_id = s.id
            WHERE s.status != 'cancelled' {start_date_str} {end_date_str}
            GROUP BY strftime('%Y-%m', s.sale_date)
            ORDER BY month
        """

        # المبيعات الشهرية
        sales_query = """
            SELECT
                strftime('%Y-%m', sale_date) as month,
                SUM(final_amount) as monthly_sales
            FROM sales
            WHERE status != 'cancelled' {start_date_str.replace("s.sale_date", "sale_date")} {end_date_str.replace("s.sale_date", "sale_date")}  # noqa: E501
            GROUP BY strftime('%Y-%m', sale_date) ORDER BY month
        """

        profit_results = self.db.fetch_all(profit_query, tuple(date_params))
        sales_results = self.db.fetch_all(sales_query, tuple(date_params))

        monthly_data = {}

        for row in sales_results:
            month = get_value(row, 'month')
            if month not in monthly_data:
                monthly_data[month] = {"sales": 0, "profit": 0}
            monthly_data[month]["sales"] = get_value(row, 'monthly_sales', 0) or 0

        for row in profit_results:
            month = get_value(row, 'month')
            if month not in monthly_data:
                monthly_data[month] = {"sales": 0, "profit": 0}
            monthly_data[month]["profit"] = get_value(row, 'monthly_profit', 0) or 0

        return monthly_data

    def get_available_reports(self) -> List[Dict[str, str]]:
        """الحصول على قائمة التقارير المتاحة"""
        return [
            {
                "type": ReportType.SALES_SUMMARY.value,
                "name": "تقرير ملخص المبيعات",
                "description": "ملخص شامل لجميع المبيعات في فترة محددة",
            },
            {
                "type": ReportType.INVENTORY_STATUS.value,
                "name": "تقرير حالة المخزون",
                "description": "حالة المخزون الحالية لجميع المنتجات",
            },
            {
                "type": ReportType.FINANCIAL_SUMMARY.value,
                "name": "تقرير الملخص المالي",
                "description": "ملخص الوضع المالي للشركة",
            },
            # تقارير المدفوعات والحسابات
            {
                "type": ReportType.PAYMENT_SUMMARY.value,
                "name": "تقرير ملخص المدفوعات",
                "description": "ملخص شامل لجميع المدفوعات والمقبوضات",
            },
            {
                "type": ReportType.RECEIVABLES_AGING.value,
                "name": "تقرير أعمار الذمم المدينة",
                "description": "تحليل أعمار المستحقات من العملاء",
            },
            {
                "type": ReportType.PAYABLES_AGING.value,
                "name": "تقرير أعمار الذمم الدائنة",
                "description": "تحليل أعمار المستحقات للموردين",
            },
            {
                "type": ReportType.CASH_FLOW.value,
                "name": "تقرير التدفق النقدي",
                "description": "تحليل التدفقات النقدية الداخلة والخارجة",
            },
            {
                "type": ReportType.PAYMENT_ANALYSIS.value,
                "name": "تقرير تحليل المدفوعات",
                "description": "تحليل مفصل لأنماط وسلوكيات المدفوعات",
            },
            {
                "type": ReportType.PAYMENT_METHODS_ANALYSIS.value,
                "name": "تقرير تحليل طرق الدفع",
                "description": "تحليل استخدام طرق الدفع المختلفة",
            },
        ]

    def get_supported_formats(self) -> List[Dict[str, Any]]:
        """الحصول على تنسيقات التصدير المدعومة"""
        formats = [
            {"format": ExportFormat.JSON.value, "name": "JSON", "available": True},
            {"format": ExportFormat.CSV.value, "name": "CSV", "available": True},
            {
                "format": ExportFormat.EXCEL.value,
                "name": "Excel",
                "available": EXCEL_AVAILABLE,
            },
            {
                "format": ExportFormat.PDF.value,
                "name": "PDF",
                "available": PDF_AVAILABLE,
            },
            {
                "format": ExportFormat.PPTX.value,
                "name": "PowerPoint",
                "available": PPTX_AVAILABLE,
            },
        ]

        return formats

    # ==================== تقارير المدفوعات والحسابات ====================

    def generate_payment_summary_report(self, filters: ReportFilter) -> ReportData:
        """توليد تقرير ملخص المدفوعات"""
        try:
            # بناء الاستعلام الأساسي
            query = """
                SELECT
                    p.payment_date,
                    p.payment_type,
                    p.payment_method,
                    p.amount,
                    p.payment_status,
                    p.currency_id,
                    p.exchange_rate,
                    p.base_amount,
                    p.converted_amount,
                    curr.code as currency_code,
                    curr.name as currency_name,
                    curr.symbol as currency_symbol,
                    CASE
                        WHEN p.payment_type = 'دفعة عميل' THEN c.name
                        WHEN p.payment_type = 'دفعة مورد' THEN s.name
                        ELSE 'غير محدد'
                    END as entity_name
                FROM payments p
                LEFT JOIN customers c ON p.customer_id = c.id AND p.payment_type = 'دفعة عميل'
                LEFT JOIN suppliers s ON p.supplier_id = s.id AND p.payment_type = 'دفعة مورد'
                LEFT JOIN currencies curr ON p.currency_id = curr.id
                WHERE 1=1
            """

            params = []

            # تطبيق الفلاتر
            if filters.start_date:
                query += " AND p.payment_date >= ?"
                params.append(filters.start_date)

            if filters.end_date:
                query += " AND p.payment_date <= ?"
                params.append(filters.end_date)

            if filters.payment_type:
                query += " AND p.payment_type = ?"
                params.append(filters.payment_type)

            if filters.payment_method:
                query += " AND p.payment_method = ?"
                params.append(filters.payment_method)

            if filters.payment_status:
                query += " AND p.payment_status = ?"
                params.append(filters.payment_status)

            if filters.entity_id:
                query += " AND p.entity_id = ?"
                params.append(filters.entity_id)

            if filters.min_amount:
                query += " AND p.amount >= ?"
                params.append(filters.min_amount)

            if filters.max_amount:
                query += " AND p.amount <= ?"
                params.append(filters.max_amount)

            query += " ORDER BY p.payment_date DESC"

            # تنفيذ الاستعلام - استخدام execute_query للحصول على dictionaries
            rows = self.db.execute_query(query, tuple(params))

            # تحويل النتائج إلى قائمة من القواميس
            data = []
            total_amount = 0
            customer_payments = 0
            supplier_payments = 0

            # تجميع حسب العملة
            currencies_summary = {}

            for row in rows:
                currency_id = row.get("currency_id")
                currency_code = row.get("currency_code") or "DZD"
                currency_name = row.get("currency_name") or "الدينار الجزائري"
                currency_symbol = row.get("currency_symbol") or "د.ج"
                base_amount = row.get("base_amount")
                converted_amount = row.get("converted_amount")
                exchange_rate = row.get("exchange_rate") or 1.0

                # استخدام base_amount إذا كان متوفراً، وإلا استخدام amount
                display_amount = float(base_amount) if base_amount else float(row.get("amount") or 0)

                payment_data = {
                    "payment_date": row.get("payment_date"),
                    "payment_type": row.get("payment_type"),
                    "payment_method": row.get("payment_method"),
                    "amount": float(row.get("amount") or 0),
                    "payment_status": row.get("payment_status"),
                    "entity_name": row.get("entity_name") or "",
                    # Multi-Currency Support
                    "currency_id": currency_id,
                    "currency_code": currency_code,
                    "currency_name": currency_name,
                    "currency_symbol": currency_symbol,
                    "base_amount": (float(base_amount) if base_amount else display_amount),
                    "converted_amount": (float(converted_amount) if converted_amount else display_amount),
                    "exchange_rate": float(exchange_rate),
                    "display_amount": display_amount,  # المبلغ للعرض (بالعملة الأساسية)
                }
                data.append(payment_data)

                # استخدام base_amount في الحسابات
                amount = display_amount
                total_amount += amount

                # تجميع حسب العملة
                if currency_code not in currencies_summary:
                    currencies_summary[currency_code] = {
                        "count": 0,
                        "amount": 0,
                        "base_amount": 0,
                    }
                currencies_summary[currency_code]["count"] += 1
                currencies_summary[currency_code]["amount"] += (
                    float(converted_amount) if converted_amount else float(row.get("amount") or 0)
                )
                currencies_summary[currency_code]["base_amount"] += amount

                if row.get("payment_type") == "دفعة عميل":
                    customer_payments += amount
                elif row.get("payment_type") == "دفعة مورد":
                    supplier_payments += amount

            # إعداد الملخص
            summary = {
                "total_payments": len(data),
                "total_amount": total_amount,  # بالعملة الأساسية
                "customer_payments": customer_payments,
                "supplier_payments": supplier_payments,
                "net_cash_flow": customer_payments - supplier_payments,
                # Multi-Currency Support
                "currencies_summary": currencies_summary,
            }

            return ReportData(
                title="تقرير ملخص المدفوعات",
                subtitle=f"من {filters.start_date or 'البداية'} إلى {filters.end_date or 'النهاية'}",
                generated_at=datetime.now(),
                filters=filters,
                data=data,
                summary=summary,
            )

        except Exception as e:
            if self.logger:
                self.logger.error(f"خطأ في توليد تقرير ملخص المدفوعات: {str(e)}")
            raise

    def generate_customer_analysis_report(self, filters: ReportFilter) -> ReportData:
        """توليد تقرير تحليل العملاء - إحصائيات الشراء لكل عميل"""
        try:
            query = """
                SELECT
                    c.id,
                    c.name,
                    c.phone,
                    COALESCE(c.email, '') as email,
                    COUNT(DISTINCT s.id) as total_orders,
                    COALESCE(SUM(s.final_amount), 0) as total_purchases,
                    COALESCE(AVG(s.final_amount), 0) as avg_order_value,
                    MAX(s.sale_date) as last_order_date
                FROM customers c
                LEFT JOIN sales s ON c.id = s.customer_id
                    AND s.is_active = 1
            """
            params = []

            if filters.start_date:
                query += " AND s.sale_date >= ?"
                params.append(filters.start_date.strftime("%Y-%m-%d"))

            if filters.end_date:
                query += " AND s.sale_date <= ?"
                params.append(filters.end_date.strftime("%Y-%m-%d"))

            if filters.customer_id:
                query += " AND c.id = ?"
                params.append(filters.customer_id)

            if filters.min_amount:
                query += " HAVING total_purchases >= ?"
                params.append(filters.min_amount)

            query += " GROUP BY c.id, c.name, c.phone, c.email ORDER BY total_purchases DESC"

            rows = self.db.execute_query(query, tuple(params))

            data = []
            total_revenue = 0
            total_orders = 0
            active_customers = 0

            for row in rows:
                total_purchases = float(row.get("total_purchases") or 0)
                total_orders_count = int(row.get("total_orders") or 0)

                if total_purchases > 0:
                    active_customers += 1

                total_revenue += total_purchases
                total_orders += total_orders_count

                data.append({
                    "customer_id": row.get("id"),
                    "customer_name": row.get("name", ""),
                    "phone": row.get("phone") or "",
                    "email": row.get("email") or "",
                    "total_orders": total_orders_count,
                    "total_purchases": total_purchases,
                    "avg_order_value": float(row.get("avg_order_value") or 0),
                    "last_order_date": row.get("last_order_date") or "",
                })

            summary = {
                "total_customers": len(data),
                "active_customers": active_customers,
                "total_revenue": total_revenue,
                "total_orders": total_orders,
                "avg_purchase_per_customer": total_revenue / len(data) if data else 0,
            }

            return ReportData(
                title="تقرير تحليل العملاء",
                subtitle=f"من {filters.start_date.strftime('%Y-%m-%d') if filters.start_date else 'البداية'} إلى {filters.end_date.strftime('%Y-%m-%d') if filters.end_date else 'النهاية'}",
                generated_at=datetime.now(),
                filters=filters,
                data=data,
                summary=summary,
            )

        except Exception as e:
            if self.logger:
                self.logger.error(f"خطأ في توليد تقرير تحليل العملاء: {str(e)}")
            raise

    def generate_supplier_analysis_report(self, filters: ReportFilter) -> ReportData:
        """توليد تقرير تحليل الموردين - إحصائيات الشراء من كل مورد"""
        try:
            query = """
                SELECT
                    s.id,
                    s.name,
                    s.phone,
                    COALESCE(s.email, '') as email,
                    COUNT(DISTINCT pur.id) as total_purchase_orders,
                    COALESCE(SUM(pur.final_amount), 0) as total_purchase_amount,
                    COALESCE(AVG(pur.final_amount), 0) as avg_order_value,
                    MAX(pur.purchase_date) as last_order_date
                FROM suppliers s
                LEFT JOIN purchases pur ON s.id = pur.supplier_id
                    AND pur.is_active = 1
            """
            params = []

            if filters.start_date:
                query += " AND pur.purchase_date >= ?"
                params.append(filters.start_date.strftime("%Y-%m-%d"))

            if filters.end_date:
                query += " AND pur.purchase_date <= ?"
                params.append(filters.end_date.strftime("%Y-%m-%d"))

            if filters.supplier_id:
                query += " AND s.id = ?"
                params.append(filters.supplier_id)

            query += " GROUP BY s.id, s.name, s.phone, s.email ORDER BY total_purchase_amount DESC"

            rows = self.db.execute_query(query, tuple(params))

            data = []
            total_spend = 0
            total_orders = 0

            for row in rows:
                total_amount = float(row.get("total_purchase_amount") or 0)
                orders_count = int(row.get("total_purchase_orders") or 0)

                total_spend += total_amount
                total_orders += orders_count

                data.append({
                    "supplier_id": row.get("id"),
                    "supplier_name": row.get("name", ""),
                    "phone": row.get("phone") or "",
                    "email": row.get("email") or "",
                    "total_purchase_orders": orders_count,
                    "total_purchase_amount": total_amount,
                    "avg_order_value": float(row.get("avg_order_value") or 0),
                    "last_order_date": row.get("last_order_date") or "",
                })

            summary = {
                "total_suppliers": len(data),
                "active_suppliers": len([d for d in data if d["total_purchase_amount"] > 0]),
                "total_spend": total_spend,
                "total_orders": total_orders,
                "avg_spend_per_supplier": total_spend / len(data) if data else 0,
            }

            return ReportData(
                title="تقرير تحليل الموردين",
                subtitle=f"من {filters.start_date.strftime('%Y-%m-%d') if filters.start_date else 'البداية'} إلى {filters.end_date.strftime('%Y-%m-%d') if filters.end_date else 'النهاية'}",
                generated_at=datetime.now(),
                filters=filters,
                data=data,
                summary=summary,
            )

        except Exception as e:
            if self.logger:
                self.logger.error(f"خطأ في توليد تقرير تحليل الموردين: {str(e)}")
            raise

    def generate_product_performance_report(self, filters: ReportFilter) -> ReportData:
        """توليد تقرير أداء المنتجات - مبيعات وكمية وإيرادات لكل منتج"""
        try:
            query = """
                SELECT
                    p.id,
                    p.name,
                    p.barcode,
                    p.unit,
                    p.cost_price,
                    p.selling_price,
                    COALESCE(c.name, '') as category_name,
                    COALESCE(SUM(si.quantity), 0) as total_qty_sold,
                    COALESCE(SUM(si.total_price), 0) as total_revenue,
                    COALESCE(SUM(si.cost_price * si.quantity), 0) as total_cost,
                    COALESCE(SUM(si.profit), 0) as total_profit,
                    COUNT(DISTINCT si.sale_id) as number_of_sales,
                    p.current_stock
                FROM products p
                LEFT JOIN categories c ON p.category_id = c.id
                LEFT JOIN sale_items si ON p.id = si.product_id
                LEFT JOIN sales s ON si.sale_id = s.id
                    AND s.is_active = 1
            """
            params = []

            if filters.start_date:
                query += " AND s.sale_date >= ?"
                params.append(filters.start_date.strftime("%Y-%m-%d"))

            if filters.end_date:
                query += " AND s.sale_date <= ?"
                params.append(filters.end_date.strftime("%Y-%m-%d"))

            if filters.category_id:
                query += " AND p.category_id = ?"
                params.append(filters.category_id)

            if filters.product_id:
                query += " AND p.id = ?"
                params.append(filters.product_id)

            query += " GROUP BY p.id, p.name, p.barcode, p.unit, p.cost_price, p.selling_price, c.name, p.current_stock"
            query += " ORDER BY total_revenue DESC"

            rows = self.db.execute_query(query, tuple(params))

            data = []
            total_revenue = 0
            total_profit = 0
            total_qty = 0

            for row in rows:
                revenue = float(row.get("total_revenue") or 0)
                profit = float(row.get("total_profit") or 0)
                qty = int(row.get("total_qty_sold") or 0)

                total_revenue += revenue
                total_profit += profit
                total_qty += qty

                cost_price = float(row.get("cost_price") or 0)
                selling_price = float(row.get("selling_price") or 0)
                total_cost = float(row.get("total_cost") or 0)

                margin_pct = (profit / revenue * 100) if revenue > 0 else 0

                data.append({
                    "product_id": row.get("id"),
                    "product_name": row.get("name", ""),
                    "barcode": row.get("barcode") or "",
                    "unit": row.get("unit") or "",
                    "category_name": row.get("category_name") or "",
                    "cost_price": cost_price,
                    "selling_price": selling_price,
                    "total_qty_sold": qty,
                    "total_revenue": revenue,
                    "total_cost": total_cost,
                    "total_profit": profit,
                    "profit_margin_pct": round(margin_pct, 2),
                    "number_of_sales": int(row.get("number_of_sales") or 0),
                    "current_stock": int(row.get("current_stock") or 0),
                })

            summary = {
                "total_products": len(data),
                "total_revenue": total_revenue,
                "total_profit": total_profit,
                "total_qty_sold": total_qty,
                "avg_margin_pct": round(
                    (total_profit / total_revenue * 100) if total_revenue > 0 else 0, 2
                ),
            }

            return ReportData(
                title="تقرير أداء المنتجات",
                subtitle=f"من {filters.start_date.strftime('%Y-%m-%d') if filters.start_date else 'البداية'} إلى {filters.end_date.strftime('%Y-%m-%d') if filters.end_date else 'النهاية'}",
                generated_at=datetime.now(),
                filters=filters,
                data=data,
                summary=summary,
            )

        except Exception as e:
            if self.logger:
                self.logger.error(f"خطأ في توليد تقرير أداء المنتجات: {str(e)}")
            raise

    def generate_stock_movement_report(self, filters: ReportFilter) -> ReportData:
        """توليد تقرير حركة المخزون - من جدول stock_movements"""
        try:
            query = """
                SELECT
                    sm.id,
                    sm.product_id,
                    COALESCE(p.name, 'غير محدد') as product_name,
                    p.barcode,
                    sm.movement_type,
                    sm.quantity,
                    sm.reference_type,
                    sm.reference_id,
                    sm.notes,
                    sm.movement_date,
                    sm.created_at,
                    COALESCE(u.username, '') as user_name
                FROM stock_movements sm
                LEFT JOIN products p ON sm.product_id = p.id
                LEFT JOIN users u ON sm.user_id = u.id
                WHERE 1=1
            """
            params = []

            if filters.start_date:
                query += " AND DATE(sm.created_at) >= ?"
                params.append(filters.start_date.strftime("%Y-%m-%d"))

            if filters.end_date:
                query += " AND DATE(sm.created_at) <= ?"
                params.append(filters.end_date.strftime("%Y-%m-%d"))

            if filters.product_id:
                query += " AND sm.product_id = ?"
                params.append(filters.product_id)

            query += " ORDER BY sm.created_at DESC"

            rows = self.db.execute_query(query, tuple(params))

            # ترجمة أنواع الحركة
            movement_type_map = {
                "in": "إدخال",
                "out": "إخراج",
                "adjustment": "تسوية",
                "return": "مرتجع",
                "transfer_in": "تحويل وارد",
                "transfer_out": "تحويل صادر",
                "إدخال": "إدخال",
                "إخراج": "إخراج",
                "تسوية": "تسوية",
                "مرتجع": "مرتجع",
            }

            data = []
            total_in = 0
            total_out = 0

            for row in rows:
                qty = int(row.get("quantity") or 0)
                mtype = str(row.get("movement_type") or "").lower()

                if mtype in ("in", "إدخال", "return", "مرتجع", "transfer_in"):
                    total_in += qty
                elif mtype in ("out", "إخراج", "transfer_out"):
                    total_out += qty

                data.append({
                    "id": row.get("id"),
                    "product_id": row.get("product_id"),
                    "product_name": row.get("product_name", ""),
                    "barcode": row.get("barcode") or "",
                    "movement_type": movement_type_map.get(
                        row.get("movement_type"), row.get("movement_type", "")
                    ),
                    "quantity": qty,
                    "reference_type": row.get("reference_type") or "",
                    "reference_id": row.get("reference_id"),
                    "notes": row.get("notes") or "",
                    "movement_date": row.get("movement_date") or "",
                    "created_at": row.get("created_at") or "",
                    "user_name": row.get("user_name") or "",
                })

            summary = {
                "total_movements": len(data),
                "total_in": total_in,
                "total_out": total_out,
                "net_movement": total_in - total_out,
            }

            return ReportData(
                title="تقرير حركة المخزون",
                subtitle=f"من {filters.start_date.strftime('%Y-%m-%d') if filters.start_date else 'البداية'} إلى {filters.end_date.strftime('%Y-%m-%d') if filters.end_date else 'النهاية'}",
                generated_at=datetime.now(),
                filters=filters,
                data=data,
                summary=summary,
            )

        except Exception as e:
            if self.logger:
                self.logger.error(f"خطأ في توليد تقرير حركة المخزون: {str(e)}")
            raise

    def generate_receivables_aging_report(self, filters: ReportFilter) -> ReportData:
        """توليد تقرير أعمار الذمم المدينة - بناءً على الفواتير غير المدفوعة"""
        try:
            from datetime import date as _date
            aging_periods = filters.aging_periods or [30, 60, 90]
            today = _date.today()

            # استعلام الفواتير غير المدفوعة مع مراعاة أعمدة due_date / paid_amount
            query = """
                SELECT
                    s.customer_id,
                    COALESCE(c.name, 'غير محدد') as customer_name,
                    COALESCE(c.phone, '') as phone,
                    COALESCE(c.email, '') as email,
                    s.id as invoice_id,
                    s.invoice_number,
                    s.sale_date,
                    s.final_amount,
                    COALESCE(s.paid_amount, 0) as paid_amount
                FROM sales s
                LEFT JOIN customers c ON s.customer_id = c.id
                WHERE s.is_active = 1
                  AND s.status NOT IN ('مدفوعة', 'paid', 'ملغية', 'cancelled',
                                       'مسودة', 'draft', 'مرتجعة', 'returned')
                  AND (s.final_amount - COALESCE(s.paid_amount, 0)) > 0
            """
            params = []

            if filters.start_date:
                query += " AND s.sale_date >= ?"
                params.append(filters.start_date.strftime("%Y-%m-%d"))

            if filters.end_date:
                query += " AND s.sale_date <= ?"
                params.append(filters.end_date.strftime("%Y-%m-%d"))

            if filters.customer_id:
                query += " AND s.customer_id = ?"
                params.append(filters.customer_id)

            # تجربة استخدام due_date إن وُجد؛ إن لم يوجد نستخدم sale_date
            try:
                test_query = "SELECT due_date FROM sales LIMIT 1"
                self.db_manager.connection.execute(test_query)
                query = query.replace(
                    "s.sale_date,",
                    "s.sale_date, COALESCE(s.due_date, s.sale_date) as ref_date,",
                )
                use_due_date = True
            except Exception:
                query = query.replace(
                    "s.sale_date,",
                    "s.sale_date, s.sale_date as ref_date,",
                )
                use_due_date = False

            rows = self.db.execute_query(query, tuple(params))

            # تجميع المبالغ حسب العميل وفترة التأخير
            buckets = {}

            for row in rows:
                customer_id = row.get("customer_id")
                remaining = float(row.get("final_amount") or 0) - float(row.get("paid_amount") or 0)

                if remaining <= 0:
                    continue

                ref_date_str = row.get("ref_date")
                if ref_date_str:
                    try:
                        if isinstance(ref_date_str, str):
                            ref_date = _date.fromisoformat(ref_date_str)
                        else:
                            ref_date = ref_date_str
                    except (ValueError, TypeError):
                        ref_date = today
                else:
                    ref_date = today

                days_overdue = (today - ref_date).days

                if days_overdue <= 0:
                    bucket_key = "current"
                elif days_overdue <= 30:
                    bucket_key = "aging_0_30"
                elif days_overdue <= 60:
                    bucket_key = "aging_31_60"
                elif days_overdue <= 90:
                    bucket_key = "aging_61_90"
                else:
                    bucket_key = "aging_over_90"

                if customer_id not in buckets:
                    buckets[customer_id] = {
                        "customer_id": customer_id,
                        "customer_name": row.get("customer_name", ""),
                        "phone": row.get("phone", ""),
                        "email": row.get("email", ""),
                        "total_sales": 0,
                        "total_paid": 0,
                        "balance": 0,
                        "current": 0,
                        "aging_0_30": 0,
                        "aging_31_60": 0,
                        "aging_61_90": 0,
                        "aging_over_90": 0,
                    }

                buckets[customer_id]["balance"] += remaining
                buckets[customer_id]["total_sales"] += float(row.get("final_amount") or 0)
                buckets[customer_id]["total_paid"] += float(row.get("paid_amount") or 0)
                buckets[customer_id][bucket_key] += remaining

            data = []
            total_balance = 0
            sum_current = 0
            sum_0_30 = 0
            sum_31_60 = 0
            sum_61_90 = 0
            sum_over_90 = 0

            for cust_id in sorted(buckets, key=lambda k: buckets[k]["balance"], reverse=True):
                entry = buckets[cust_id]
                if not filters.include_zero_balances and entry["balance"] <= 0:
                    continue
                data.append(entry)
                total_balance += entry["balance"]
                sum_current += entry["current"]
                sum_0_30 += entry["aging_0_30"]
                sum_31_60 += entry["aging_31_60"]
                sum_61_90 += entry["aging_61_90"]
                sum_over_90 += entry["aging_over_90"]

            summary = {
                "total_customers": len(data),
                "total_balance": total_balance,
                "aging_periods": aging_periods,
                "sum_current": sum_current,
                "sum_0_30": sum_0_30,
                "sum_31_60": sum_31_60,
                "sum_61_90": sum_61_90,
                "sum_over_90": sum_over_90,
            }

            return ReportData(
                title="تقرير أعمار الذمم المدينة",
                subtitle=f"كما في {today.isoformat()}",
                generated_at=datetime.now(),
                filters=filters,
                data=data,
                summary=summary,
            )

        except Exception as e:
            if self.logger:
                self.logger.error(f"خطأ في توليد تقرير أعمار الذمم المدينة: {str(e)}")
            raise

    def generate_payables_aging_report(self, filters: ReportFilter) -> ReportData:
        """توليد تقرير أعمار الذمم الدائنة - بناءً على فواتير المشتريات غير المدفوعة"""
        try:
            from datetime import date as _date
            aging_periods = filters.aging_periods or [30, 60, 90]
            today = _date.today()

            # استعلام فواتير المشتريات غير المدفوعة
            query = """
                SELECT
                    p.supplier_id,
                    COALESCE(sup.name, 'غير محدد') as supplier_name,
                    COALESCE(sup.phone, '') as phone,
                    COALESCE(sup.email, '') as email,
                    p.id as invoice_id,
                    p.invoice_number,
                    p.purchase_date,
                    COALESCE(p.final_amount, p.total_amount) as final_amount,
                    COALESCE(p.paid_amount, 0) as paid_amount
                FROM purchases p
                LEFT JOIN suppliers sup ON p.supplier_id = sup.id
                WHERE p.is_active = 1
                  AND p.status NOT IN ('مدفوعة', 'paid', 'ملغية', 'cancelled', 'مرتجعة', 'returned')
                  AND (COALESCE(p.final_amount, p.total_amount) - COALESCE(p.paid_amount, 0)) > 0
            """
            params = []

            if filters.start_date:
                query += " AND p.purchase_date >= ?"
                params.append(filters.start_date.strftime("%Y-%m-%d"))

            if filters.end_date:
                query += " AND p.purchase_date <= ?"
                params.append(filters.end_date.strftime("%Y-%m-%d"))

            if filters.supplier_id:
                query += " AND p.supplier_id = ?"
                params.append(filters.supplier_id)

            # تجربة استخدام due_date إن وُجد
            try:
                test_query = "SELECT due_date FROM purchases LIMIT 1"
                self.db_manager.connection.execute(test_query)
                query = query.replace(
                    "p.purchase_date,",
                    "p.purchase_date, COALESCE(p.due_date, p.purchase_date) as ref_date,",
                )
            except Exception:
                query = query.replace(
                    "p.purchase_date,",
                    "p.purchase_date, p.purchase_date as ref_date,",
                )

            rows = self.db.execute_query(query, tuple(params))

            # تجميع المبالغ حسب المورد وفترة التأخير
            buckets = {}

            for row in rows:
                supplier_id = row.get("supplier_id")
                final_amount = float(row.get("final_amount") or 0)
                paid_amount = float(row.get("paid_amount") or 0)
                remaining = final_amount - paid_amount

                if remaining <= 0:
                    continue

                ref_date_str = row.get("ref_date")
                if ref_date_str:
                    try:
                        if isinstance(ref_date_str, str):
                            ref_date = _date.fromisoformat(ref_date_str)
                        else:
                            ref_date = ref_date_str
                    except (ValueError, TypeError):
                        ref_date = today
                else:
                    ref_date = today

                days_overdue = (today - ref_date).days

                if days_overdue <= 0:
                    bucket_key = "current"
                elif days_overdue <= 30:
                    bucket_key = "aging_0_30"
                elif days_overdue <= 60:
                    bucket_key = "aging_31_60"
                elif days_overdue <= 90:
                    bucket_key = "aging_61_90"
                else:
                    bucket_key = "aging_over_90"

                if supplier_id not in buckets:
                    buckets[supplier_id] = {
                        "supplier_id": supplier_id,
                        "supplier_name": row.get("supplier_name", ""),
                        "phone": row.get("phone", ""),
                        "email": row.get("email", ""),
                        "total_purchases": 0,
                        "total_paid": 0,
                        "balance": 0,
                        "current": 0,
                        "aging_0_30": 0,
                        "aging_31_60": 0,
                        "aging_61_90": 0,
                        "aging_over_90": 0,
                    }

                buckets[supplier_id]["balance"] += remaining
                buckets[supplier_id]["total_purchases"] += final_amount
                buckets[supplier_id]["total_paid"] += paid_amount
                buckets[supplier_id][bucket_key] += remaining

            data = []
            total_balance = 0
            sum_current = 0
            sum_0_30 = 0
            sum_31_60 = 0
            sum_61_90 = 0
            sum_over_90 = 0

            for sup_id in sorted(buckets, key=lambda k: buckets[k]["balance"], reverse=True):
                entry = buckets[sup_id]
                if not filters.include_zero_balances and entry["balance"] <= 0:
                    continue
                data.append(entry)
                total_balance += entry["balance"]
                sum_current += entry["current"]
                sum_0_30 += entry["aging_0_30"]
                sum_31_60 += entry["aging_31_60"]
                sum_61_90 += entry["aging_61_90"]
                sum_over_90 += entry["aging_over_90"]

            summary = {
                "total_suppliers": len(data),
                "total_balance": total_balance,
                "aging_periods": aging_periods,
                "sum_current": sum_current,
                "sum_0_30": sum_0_30,
                "sum_31_60": sum_31_60,
                "sum_61_90": sum_61_90,
                "sum_over_90": sum_over_90,
            }

            return ReportData(
                title="تقرير أعمار الذمم الدائنة",
                subtitle=f"كما في {today.isoformat()}",
                generated_at=datetime.now(),
                filters=filters,
                data=data,
                summary=summary,
            )

        except Exception as e:
            if self.logger:
                self.logger.error(f"خطأ في توليد تقرير أعمار الذمم الدائنة: {str(e)}")
            raise

    def generate_cash_flow_report(self, filters: ReportFilter) -> ReportData:
        """توليد تقرير التدفق النقدي"""
        try:
            query = """
                SELECT
                    DATE(p.payment_date) as payment_date,
                    p.payment_type,
                    p.payment_method,
                    SUM(COALESCE(p.base_amount, p.amount)) as daily_amount
                FROM payments p
                WHERE p.payment_status = 'مكتمل'
            """

            params = []

            if filters.start_date:
                query += " AND p.payment_date >= ?"
                params.append(filters.start_date)

            if filters.end_date:
                query += " AND p.payment_date <= ?"
                params.append(filters.end_date)

            if filters.payment_method:
                query += " AND p.payment_method = ?"
                params.append(filters.payment_method)

            query += " GROUP BY DATE(p.payment_date), p.payment_type, p.payment_method"
            query += " ORDER BY payment_date DESC"

            cursor = self.db_manager.connection.cursor()
            cursor.execute(query, params)
            results = cursor.fetchall()

            # تجميع البيانات حسب التاريخ
            daily_flow = {}
            total_inflow = 0
            total_outflow = 0

            for row in results:
                date = get_value(row, 'payment_date')
                payment_type = get_value(row, 'payment_type')
                payment_method = get_value(row, 'payment_method')
                amount = float(get_value(row, 'daily_amount') or 0)

                if date not in daily_flow:
                    daily_flow[date] = {
                        "date": date,
                        "inflow": 0,
                        "outflow": 0,
                        "net": 0,
                        "methods": {},
                    }

                if payment_method not in daily_flow[date]["methods"]:
                    daily_flow[date]["methods"][payment_method] = {
                        "inflow": 0,
                        "outflow": 0,
                    }

                if payment_type == "دفعة عميل":
                    daily_flow[date]["inflow"] += amount
                    daily_flow[date]["methods"][payment_method]["inflow"] += amount
                    total_inflow += amount
                elif payment_type == "دفعة مورد":
                    daily_flow[date]["outflow"] += amount
                    daily_flow[date]["methods"][payment_method]["outflow"] += amount
                    total_outflow += amount

                daily_flow[date]["net"] = daily_flow[date]["inflow"] - daily_flow[date]["outflow"]

            # تحويل إلى قائمة
            data = list(daily_flow.values())

            summary = {
                "total_inflow": total_inflow,
                "total_outflow": total_outflow,
                "net_cash_flow": total_inflow - total_outflow,
                "period_days": len(data),
            }

            return ReportData(
                title="تقرير التدفق النقدي",
                subtitle=f"من {filters.start_date or 'البداية'} إلى {filters.end_date or 'النهاية'}",
                generated_at=datetime.now(),
                filters=filters,
                data=data,
                summary=summary,
            )

        except Exception as e:
            if self.logger:
                self.logger.error(f"خطأ في توليد تقرير التدفق النقدي: {str(e)}")
            raise

    def generate_payment_analysis_report(self, filters: ReportFilter) -> ReportData:
        """توليد تقرير تحليل المدفوعات"""
        try:
            # تحليل المدفوعات حسب النوع والطريقة
            query = """
                SELECT
                    p.payment_type,
                    p.payment_method,
                    COUNT(*) as transaction_count,
                    SUM(p.amount) as total_amount,
                    AVG(p.amount) as average_amount,
                    MIN(p.amount) as min_amount,
                    MAX(p.amount) as max_amount
                FROM payments p
                WHERE p.payment_status = 'مكتمل'
            """

            params = []

            if filters.start_date:
                query += " AND p.payment_date >= ?"
                params.append(filters.start_date)

            if filters.end_date:
                query += " AND p.payment_date <= ?"
                params.append(filters.end_date)

            query += " GROUP BY p.payment_type, p.payment_method"
            query += " ORDER BY total_amount DESC"

            cursor = self.db_manager.connection.cursor()
            cursor.execute(query, params)
            results = cursor.fetchall()

            data = []
            total_transactions = 0
            total_amount = 0

            for row in results:
                transaction_count = get_value(row, 'transaction_count', 0)
                total_amt = float(get_value(row, 'total_amount') or 0)
                analysis_data = {
                    "payment_type": get_value(row, 'payment_type'),
                    "payment_method": get_value(row, 'payment_method'),
                    "transaction_count": transaction_count,
                    "total_amount": total_amt,
                    "average_amount": float(get_value(row, 'average_amount') or 0),
                    "min_amount": float(get_value(row, 'min_amount') or 0),
                    "max_amount": float(get_value(row, 'max_amount') or 0),
                    "percentage": 0,  # سيتم حسابها لاحقاً
                }
                data.append(analysis_data)
                total_transactions += transaction_count
                total_amount += total_amt

            # حساب النسب المئوية
            for item in data:
                if total_amount > 0:
                    item["percentage"] = (item["total_amount"] / total_amount) * 100

            summary = {
                "total_transactions": total_transactions,
                "total_amount": total_amount,
                "average_transaction": (total_amount / total_transactions if total_transactions > 0 else 0),
            }

            return ReportData(
                title="تقرير تحليل المدفوعات",
                subtitle=f"من {filters.start_date or 'البداية'} إلى {filters.end_date or 'النهاية'}",
                generated_at=datetime.now(),
                filters=filters,
                data=data,
                summary=summary,
            )

        except Exception as e:
            if self.logger:
                self.logger.error(f"خطأ في توليد تقرير تحليل المدفوعات: {str(e)}")
            raise

    def generate_payment_methods_analysis_report(self, filters: ReportFilter) -> ReportData:
        """توليد تقرير تحليل طرق الدفع"""
        try:
            query = """
                SELECT
                    p.payment_method,
                    COUNT(*) as usage_count,
                    SUM(p.amount) as total_amount,
                    AVG(p.amount) as average_amount,
                    COUNT(CASE WHEN p.payment_type = 'دفعة عميل' THEN 1 END) as customer_payments,
                    COUNT(CASE WHEN p.payment_type = 'دفعة مورد' THEN 1 END) as supplier_payments
                FROM payments p
                WHERE p.payment_status = 'مكتمل'
            """

            params = []

            if filters.start_date:
                query += " AND p.payment_date >= ?"
                params.append(filters.start_date)

            if filters.end_date:
                query += " AND p.payment_date <= ?"
                params.append(filters.end_date)

            query += " GROUP BY p.payment_method"
            query += " ORDER BY total_amount DESC"

            cursor = self.db_manager.connection.cursor()
            cursor.execute(query, params)
            results = cursor.fetchall()

            data = []
            total_usage = 0
            total_amount = 0

            for row in results:
                usage_count = get_value(row, 'usage_count', 0)
                total_amt = float(get_value(row, 'total_amount') or 0)
                method_data = {
                    "payment_method": get_value(row, 'payment_method'),
                    "usage_count": usage_count,
                    "total_amount": total_amt,
                    "average_amount": float(get_value(row, 'average_amount') or 0),
                    "customer_payments": get_value(row, 'customer_payments'),
                    "supplier_payments": get_value(row, 'supplier_payments'),
                    "usage_percentage": 0,
                    "amount_percentage": 0,
                }
                data.append(method_data)
                total_usage += usage_count
                total_amount += total_amt

            # حساب النسب المئوية
            for item in data:
                if total_usage > 0:
                    item["usage_percentage"] = (item["usage_count"] / total_usage) * 100
                if total_amount > 0:
                    item["amount_percentage"] = (item["total_amount"] / total_amount) * 100

            summary = {
                "total_methods": len(data),
                "total_usage": total_usage,
                "total_amount": total_amount,
                "most_used_method": data[0]["payment_method"] if data else None,
                "highest_amount_method": (
                    max(data, key=lambda x: x["total_amount"])["payment_method"] if data else None
                ),
            }

            return ReportData(
                title="تقرير تحليل طرق الدفع",
                subtitle=f"من {filters.start_date or 'البداية'} إلى {filters.end_date or 'النهاية'}",
                generated_at=datetime.now(),
                filters=filters,
                data=data,
                summary=summary,
            )

        except Exception as e:
            if self.logger:
                self.logger.error(f"خطأ في توليد تقرير تحليل طرق الدفع: {str(e)}")
            raise

    # ===== Multi-Warehouse Reports =====

    def generate_warehouse_inventory_report(self, filters: ReportFilter) -> ReportData:
        """توليد تقرير المخزون حسب المستودع"""
        try:
            from src.services.warehouse_service import WarehouseService

            warehouse_service = WarehouseService(self.db_manager)

            # تحديد المستودع
            warehouse_id = filters.warehouse_id
            warehouses = []

            if warehouse_id:
                # مستودع محدد
                warehouse = warehouse_service.get_warehouse(warehouse_id)
                if warehouse:
                    warehouses = [warehouse]
            else:
                # جميع المستودعات
                warehouses = warehouse_service.get_all_warehouses(include_inactive=False)

            data = []
            total_stock_value = 0
            total_products = 0

            for warehouse in warehouses:
                inventory = warehouse_service.get_warehouse_inventory(warehouse.id)

                for inv in inventory:
                    stock_value = inv.quantity * (inv.cost_price or 0)
                    total_stock_value += stock_value
                    total_products += 1

                    data.append(
                        {
                            "warehouse_name": warehouse.name,
                            "warehouse_code": warehouse.code,
                            "product_name": inv.product_name or "",
                            "product_id": inv.product_id,
                            "quantity": inv.quantity,
                            "reserved_quantity": inv.reserved_quantity,
                            "available_quantity": inv.available_quantity,
                            "min_stock": inv.min_stock,
                            "reorder_point": inv.reorder_point,
                            "cost_price": inv.cost_price or 0,
                            "stock_value": stock_value,
                            "status": (
                                "نفد"
                                if inv.available_quantity <= 0
                                else ("منخفض" if inv.available_quantity <= inv.min_stock else "جيد")
                            ),
                        }
                    )

            # تجميع حسب المستودع
            warehouses_summary = {}
            for item in data:
                wh_name = item["warehouse_name"]
                if wh_name not in warehouses_summary:
                    warehouses_summary[wh_name] = {
                        "products_count": 0,
                        "total_value": 0,
                        "low_stock_count": 0,
                        "out_of_stock_count": 0,
                    }
                warehouses_summary[wh_name]["products_count"] += 1
                warehouses_summary[wh_name]["total_value"] += item["stock_value"]
                if item["status"] == "منخفض":
                    warehouses_summary[wh_name]["low_stock_count"] += 1
                elif item["status"] == "نفد":
                    warehouses_summary[wh_name]["out_of_stock_count"] += 1

            summary = {
                "total_warehouses": len(warehouses),
                "total_products": total_products,
                "total_stock_value": total_stock_value,
                "warehouses_summary": warehouses_summary,
            }

            return ReportData(
                title="تقرير المخزون حسب المستودع",
                subtitle=(
                    f"{warehouses[0].name if len(warehouses) == 1 else 'جميع المستودعات'}"
                    if warehouses
                    else "لا توجد مستودعات"
                ),
                generated_at=datetime.now(),
                filters=filters,
                data=data,
                summary=summary,
            )

        except Exception as e:
            if self.logger:
                self.logger.error(f"خطأ في توليد تقرير مخزون المستودع: {str(e)}")
            raise

    def generate_warehouse_transfers_report(self, filters: ReportFilter) -> ReportData:
        """توليد تقرير حركات النقل بين المستودعات"""
        try:
            from src.services.warehouse_service import WarehouseService

            warehouse_service = WarehouseService(self.db_manager)

            # الحصول على التحويلات
            warehouse_id = filters.warehouse_id
            transfers = warehouse_service.get_transfers(warehouse_id=warehouse_id)

            # فلترة حسب التاريخ
            if filters.start_date or filters.end_date:
                filtered_transfers = []
                for transfer in transfers:
                    if transfer.transfer_date:
                        transfer_date = (
                            transfer.transfer_date
                            if isinstance(transfer.transfer_date, datetime)
                            else datetime.combine(transfer.transfer_date, datetime.min.time())
                        )
                        if filters.start_date and transfer_date < filters.start_date:
                            continue
                        if filters.end_date and transfer_date > filters.end_date:
                            continue
                    filtered_transfers.append(transfer)
                transfers = filtered_transfers

            data = []
            total_quantity = 0
            completed_count = 0
            pending_count = 0

            for transfer in transfers:
                total_quantity += transfer.quantity
                if transfer.status == "completed":
                    completed_count += 1
                elif transfer.status == "pending":
                    pending_count += 1

                data.append(
                    {
                        "transfer_number": transfer.transfer_number,
                        "from_warehouse": transfer.from_warehouse_name or "",
                        "to_warehouse": transfer.to_warehouse_name or "",
                        "product_name": transfer.product_name or "",
                        "quantity": transfer.quantity,
                        "status": transfer.status,
                        "transfer_date": (
                            transfer.transfer_date.strftime("%Y-%m-%d %H:%M") if transfer.transfer_date else ""
                        ),
                        "received_date": (
                            transfer.received_date.strftime("%Y-%m-%d %H:%M") if transfer.received_date else ""
                        ),
                        "notes": transfer.notes or "",
                    }
                )

            summary = {
                "total_transfers": len(transfers),
                "total_quantity": total_quantity,
                "completed_count": completed_count,
                "pending_count": pending_count,
                "in_transit_count": len([t for t in transfers if t.status == "in_transit"]),
                "cancelled_count": len([t for t in transfers if t.status == "cancelled"]),
            }

            return ReportData(
                title="تقرير حركات النقل بين المستودعات",
                subtitle=f"من {filters.start_date.strftime('%Y-%m-%d') if filters.start_date else 'البداية'} إلى {filters.end_date.strftime('%Y-%m-%d') if filters.end_date else 'النهاية'}",  # noqa: E501
                generated_at=datetime.now(),
                filters=filters,
                data=data,
                summary=summary,
            )

        except Exception as e:
            if self.logger:
                self.logger.error(f"خطأ في توليد تقرير حركات النقل: {str(e)}")
            raise

    def generate_warehouse_low_stock_report(self, filters: ReportFilter) -> ReportData:
        """توليد تقرير المنتجات منخفضة المخزون حسب المستودع"""
        try:
            from src.services.warehouse_service import WarehouseService

            warehouse_service = WarehouseService(self.db_manager)

            warehouse_id = filters.warehouse_id
            warehouses = []

            if warehouse_id:
                warehouse = warehouse_service.get_warehouse(warehouse_id)
                if warehouse:
                    warehouses = [warehouse]
            else:
                warehouses = warehouse_service.get_all_warehouses(include_inactive=False)

            data = []

            for warehouse in warehouses:
                inventory = warehouse_service.get_warehouse_inventory(warehouse.id)

                for inv in inventory:
                    # فقط المنتجات منخفضة أو نافذة المخزون
                    if inv.available_quantity <= inv.min_stock:
                        stock_value = inv.quantity * (inv.cost_price or 0)

                        data.append(
                            {
                                "warehouse_name": warehouse.name,
                                "warehouse_code": warehouse.code,
                                "product_name": inv.product_name or "",
                                "product_id": inv.product_id,
                                "current_stock": inv.quantity,
                                "available_quantity": inv.available_quantity,
                                "min_stock": inv.min_stock,
                                "reorder_point": inv.reorder_point,
                                "needed_quantity": max(0, inv.reorder_point - inv.available_quantity),
                                "cost_price": inv.cost_price or 0,
                                "stock_value": stock_value,
                                "status": ("نفد" if inv.available_quantity <= 0 else "منخفض"),
                            }
                        )

            # ترتيب حسب الأولوية (نفد أولاً، ثم الأكثر حاجة)
            data.sort(key=lambda x: (0 if x["status"] == "نفد" else 1, -x["needed_quantity"]))

            # تجميع حسب المستودع
            warehouses_summary = {}
            for item in data:
                wh_name = item["warehouse_name"]
                if wh_name not in warehouses_summary:
                    warehouses_summary[wh_name] = {
                        "low_stock_count": 0,
                        "out_of_stock_count": 0,
                        "total_needed_value": 0,
                    }
                if item["status"] == "نفد":
                    warehouses_summary[wh_name]["out_of_stock_count"] += 1
                else:
                    warehouses_summary[wh_name]["low_stock_count"] += 1
                warehouses_summary[wh_name]["total_needed_value"] += item["needed_quantity"] * item["cost_price"]

            summary = {
                "total_items": len(data),
                "out_of_stock_count": len([d for d in data if d["status"] == "نفد"]),
                "low_stock_count": len([d for d in data if d["status"] == "منخفض"]),
                "warehouses_summary": warehouses_summary,
            }

            return ReportData(
                title="تقرير المنتجات منخفضة المخزون",
                subtitle=f"حسب المستودعات - {len(warehouses)} مستودع",
                generated_at=datetime.now(),
                filters=filters,
                data=data,
                summary=summary,
            )

        except Exception as e:
            if self.logger:
                self.logger.error(f"خطأ في توليد تقرير المنتجات منخفضة المخزون: {str(e)}")
            raise

    def generate_warehouse_summary_report(self, filters: ReportFilter) -> ReportData:
        """توليد تقرير ملخص المستودعات"""
        try:
            from src.services.warehouse_service import WarehouseService

            warehouse_service = WarehouseService(self.db_manager)

            warehouses = warehouse_service.get_all_warehouses(include_inactive=False)

            data = []
            total_stock_value = 0
            total_products = 0

            for warehouse in warehouses:
                summary = warehouse_service.get_warehouse_summary(warehouse.id)

                total_stock_value += summary.get("total_stock_value", 0)
                total_products += summary.get("total_products", 0)

                data.append(
                    {
                        "warehouse_name": warehouse.name,
                        "warehouse_code": warehouse.code,
                        "city": warehouse.city or "",
                        "manager_name": warehouse.manager_name or "",
                        "phone": warehouse.phone or "",
                        "is_active": warehouse.is_active,
                        "is_default": warehouse.is_default,
                        "total_products": summary.get("total_products", 0),
                        "total_stock_value": summary.get("total_stock_value", 0),
                        "low_stock_count": summary.get("low_stock_count", 0),
                        "out_of_stock_count": summary.get("out_of_stock_count", 0),
                    }
                )

            summary_data = {
                "total_warehouses": len(warehouses),
                "active_warehouses": len([w for w in warehouses if w.is_active]),
                "total_products": total_products,
                "total_stock_value": total_stock_value,
                "average_stock_per_warehouse": (total_stock_value / len(warehouses) if warehouses else 0),
            }

            return ReportData(
                title="تقرير ملخص المستودعات",
                subtitle="نظرة شاملة على جميع المستودعات",
                generated_at=datetime.now(),
                filters=filters,
                data=data,
                summary=summary_data,
            )

        except Exception as e:
            if self.logger:
                self.logger.error(f"خطأ في توليد تقرير ملخص المستودعات: {str(e)}")
            raise
