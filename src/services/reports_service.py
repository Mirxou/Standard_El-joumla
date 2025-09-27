#!/usr/bin/env python3
"""
خدمة التقارير - Reports Service
تدير إنشاء وتصدير التقارير المختلفة للنظام
"""

import os
import json
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, asdict
from enum import Enum
import logging

# استيراد المكتبات المطلوبة للتصدير
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from weasyprint import HTML, CSS
    from jinja2 import Template
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

from ..core.database_manager import DatabaseManager
from ..utils.logger import setup_logger, DatabaseLogger
from ..models.product import ProductManager
from ..models.sale import SaleManager
from ..models.purchase import PurchaseManager
from ..models.customer import CustomerManager
from ..models.supplier import SupplierManager


class ReportType(Enum):
    """أنواع التقارير المتاحة"""
    SALES_SUMMARY = "sales_summary"
    INVENTORY_STATUS = "inventory_status"
    FINANCIAL_SUMMARY = "financial_summary"
    CUSTOMER_ANALYSIS = "customer_analysis"
    SUPPLIER_ANALYSIS = "supplier_analysis"
    PRODUCT_PERFORMANCE = "product_performance"
    PROFIT_LOSS = "profit_loss"
    STOCK_MOVEMENT = "stock_movement"
    # تقارير المدفوعات والحسابات
    PAYMENT_SUMMARY = "payment_summary"
    RECEIVABLES_AGING = "receivables_aging"
    PAYABLES_AGING = "payables_aging"
    CASH_FLOW = "cash_flow"
    PAYMENT_ANALYSIS = "payment_analysis"
    PAYMENT_METHODS_ANALYSIS = "payment_methods_analysis"


class ExportFormat(Enum):
    """تنسيقات التصدير المتاحة"""
    PDF = "pdf"
    EXCEL = "xlsx"
    JSON = "json"
    CSV = "csv"


@dataclass
class ReportFilter:
    """فلاتر التقارير"""
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None
    category_id: Optional[int] = None
    customer_id: Optional[int] = None
    supplier_id: Optional[int] = None
    product_id: Optional[int] = None
    user_id: Optional[int] = None
    min_amount: Optional[float] = None
    max_amount: Optional[float] = None
    # فلاتر خاصة بتقارير المدفوعات
    payment_type: Optional[str] = None
    payment_method: Optional[str] = None
    payment_status: Optional[str] = None
    account_type: Optional[str] = None
    entity_id: Optional[int] = None
    aging_periods: Optional[List[int]] = None  # فترات الأعمار [30, 60, 90]
    include_zero_balances: bool = True
    group_by: Optional[str] = None  # تجميع البيانات


@dataclass
class ReportData:
    """بيانات التقرير"""
    title: str
    subtitle: str
    generated_at: datetime
    filters: ReportFilter
    data: List[Dict[str, Any]]
    summary: Dict[str, Any]
    charts_data: Optional[Dict[str, Any]] = None


class ReportsService:
    """خدمة التقارير"""
    
    def __init__(self, db_manager: DatabaseManager):
        self.db = db_manager
        self.logger = setup_logger(__name__)
        self.db_logger = DatabaseLogger(db_manager)
        
        # إنشاء مديري النماذج
        self.product_manager = ProductManager(db_manager)
        self.sale_manager = SaleManager(db_manager)
        self.purchase_manager = PurchaseManager(db_manager)
        self.customer_manager = CustomerManager(db_manager)
        self.supplier_manager = SupplierManager(db_manager)
        
        # إعداد مجلد التقارير
        self.reports_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'data', 'exports')
        os.makedirs(self.reports_dir, exist_ok=True)
    
    def generate_sales_summary_report(self, filters: ReportFilter) -> ReportData:
        """إنشاء تقرير ملخص المبيعات"""
        try:
            # بناء استعلام المبيعات
            query = """
                SELECT 
                    s.id,
                    s.invoice_number,
                    s.sale_date,
                    s.total_amount,
                    s.discount_amount,
                    s.tax_amount,
                    s.final_amount,
                    s.payment_method,
                    s.status,
                    c.name as customer_name,
                    u.username as user_name
                FROM sales s
                LEFT JOIN customers c ON s.customer_id = c.id
                LEFT JOIN users u ON s.user_id = u.id
                WHERE 1=1
            """
            
            params = []
            
            # تطبيق الفلاتر
            if filters.start_date:
                query += " AND s.sale_date >= ?"
                params.append(filters.start_date.strftime('%Y-%m-%d'))
            
            if filters.end_date:
                query += " AND s.sale_date <= ?"
                params.append(filters.end_date.strftime('%Y-%m-%d'))
            
            if filters.customer_id:
                query += " AND s.customer_id = ?"
                params.append(filters.customer_id)
            
            if filters.user_id:
                query += " AND s.user_id = ?"
                params.append(filters.user_id)
            
            if filters.min_amount:
                query += " AND s.final_amount >= ?"
                params.append(filters.min_amount)
            
            if filters.max_amount:
                query += " AND s.final_amount <= ?"
                params.append(filters.max_amount)
            
            query += " ORDER BY s.sale_date DESC"
            
            # تنفيذ الاستعلام
            sales_data = self.db.execute_query(query, params)
            
            # حساب الملخص
            total_sales = len(sales_data)
            total_amount = sum(row['final_amount'] for row in sales_data)
            total_discount = sum(row['discount_amount'] or 0 for row in sales_data)
            total_tax = sum(row['tax_amount'] or 0 for row in sales_data)
            
            # تجميع البيانات حسب طريقة الدفع
            payment_methods = {}
            for row in sales_data:
                method = row['payment_method']
                if method not in payment_methods:
                    payment_methods[method] = {'count': 0, 'amount': 0}
                payment_methods[method]['count'] += 1
                payment_methods[method]['amount'] += row['final_amount']
            
            summary = {
                'total_sales': total_sales,
                'total_amount': total_amount,
                'total_discount': total_discount,
                'total_tax': total_tax,
                'average_sale': total_amount / total_sales if total_sales > 0 else 0,
                'payment_methods': payment_methods
            }
            
            # بيانات الرسوم البيانية
            charts_data = {
                'daily_sales': self._get_daily_sales_chart_data(filters),
                'payment_methods': payment_methods,
                'top_customers': self._get_top_customers_chart_data(filters)
            }
            
            return ReportData(
                title="تقرير ملخص المبيعات",
                subtitle=f"من {filters.start_date.strftime('%Y-%m-%d') if filters.start_date else 'البداية'} إلى {filters.end_date.strftime('%Y-%m-%d') if filters.end_date else 'النهاية'}",
                generated_at=datetime.now(),
                filters=filters,
                data=sales_data,
                summary=summary,
                charts_data=charts_data
            )
            
        except Exception as e:
            self.logger.error(f"خطأ في إنشاء تقرير ملخص المبيعات: {e}")
            raise
    
    def generate_inventory_status_report(self, filters: ReportFilter) -> ReportData:
        """إنشاء تقرير حالة المخزون"""
        try:
            query = """
                SELECT 
                    p.id,
                    p.name,
                    p.barcode,
                    p.current_stock,
                    p.min_stock_level,
                    p.max_stock_level,
                    p.cost_price,
                    p.selling_price,
                    p.current_stock * p.cost_price as stock_value,
                    c.name as category_name,
                    s.name as supplier_name,
                    CASE 
                        WHEN p.current_stock <= 0 THEN 'نفد المخزون'
                        WHEN p.current_stock <= p.min_stock_level THEN 'مخزون منخفض'
                        WHEN p.current_stock >= p.max_stock_level THEN 'مخزون مرتفع'
                        ELSE 'طبيعي'
                    END as stock_status
                FROM products p
                LEFT JOIN categories c ON p.category_id = c.id
                LEFT JOIN suppliers s ON p.supplier_id = s.id
                WHERE p.is_active = 1
            """
            
            params = []
            
            if filters.category_id:
                query += " AND p.category_id = ?"
                params.append(filters.category_id)
            
            if filters.supplier_id:
                query += " AND p.supplier_id = ?"
                params.append(filters.supplier_id)
            
            query += " ORDER BY p.name"
            
            products_data = self.db.execute_query(query, params)
            
            # حساب الملخص
            total_products = len(products_data)
            total_stock_value = sum(row['stock_value'] for row in products_data)
            out_of_stock = len([p for p in products_data if p['current_stock'] <= 0])
            low_stock = len([p for p in products_data if 0 < p['current_stock'] <= p['min_stock_level']])
            high_stock = len([p for p in products_data if p['current_stock'] >= p['max_stock_level']])
            
            # تجميع حسب الفئات
            categories_summary = {}
            for row in products_data:
                cat = row['category_name'] or 'بدون فئة'
                if cat not in categories_summary:
                    categories_summary[cat] = {'count': 0, 'value': 0}
                categories_summary[cat]['count'] += 1
                categories_summary[cat]['value'] += row['stock_value']
            
            summary = {
                'total_products': total_products,
                'total_stock_value': total_stock_value,
                'out_of_stock': out_of_stock,
                'low_stock': low_stock,
                'high_stock': high_stock,
                'normal_stock': total_products - out_of_stock - low_stock - high_stock,
                'categories_summary': categories_summary
            }
            
            charts_data = {
                'stock_status': {
                    'نفد المخزون': out_of_stock,
                    'مخزون منخفض': low_stock,
                    'مخزون مرتفع': high_stock,
                    'طبيعي': summary['normal_stock']
                },
                'categories_value': categories_summary
            }
            
            return ReportData(
                title="تقرير حالة المخزون",
                subtitle="حالة المخزون الحالية لجميع المنتجات",
                generated_at=datetime.now(),
                filters=filters,
                data=products_data,
                summary=summary,
                charts_data=charts_data
            )
            
        except Exception as e:
            self.logger.error(f"خطأ في إنشاء تقرير حالة المخزون: {e}")
            raise
    
    def generate_financial_summary_report(self, filters: ReportFilter) -> ReportData:
        """إنشاء تقرير الملخص المالي"""
        try:
            # المبيعات
            sales_query = """
                SELECT 
                    SUM(final_amount) as total_sales,
                    SUM(discount_amount) as total_discounts,
                    SUM(tax_amount) as total_tax,
                    COUNT(*) as sales_count
                FROM sales 
                WHERE status != 'cancelled'
            """
            
            # المشتريات
            purchases_query = """
                SELECT 
                    SUM(total_amount) as total_purchases,
                    COUNT(*) as purchases_count
                FROM purchases 
                WHERE status != 'cancelled'
            """
            
            params = []
            
            if filters.start_date:
                sales_query += " AND sale_date >= ?"
                purchases_query += " AND purchase_date >= ?"
                params.extend([filters.start_date.strftime('%Y-%m-%d')] * 2)
            
            if filters.end_date:
                sales_query += " AND sale_date <= ?"
                purchases_query += " AND purchase_date <= ?"
                params.extend([filters.end_date.strftime('%Y-%m-%d')] * 2)
            
            # تنفيذ الاستعلامات
            sales_result = self.db.execute_query(sales_query, params[:len(params)//2] if params else [])
            purchases_result = self.db.execute_query(purchases_query, params[len(params)//2:] if params else [])
            
            sales_data = sales_result[0] if sales_result else {}
            purchases_data = purchases_result[0] if purchases_result else {}
            
            # حساب الأرباح
            total_sales = sales_data.get('total_sales', 0) or 0
            total_purchases = purchases_data.get('total_purchases', 0) or 0
            gross_profit = total_sales - total_purchases
            
            # حساب قيمة المخزون الحالي
            inventory_query = "SELECT SUM(current_stock * cost_price) as inventory_value FROM products WHERE is_active = 1"
            inventory_result = self.db.execute_query(inventory_query)
            inventory_value = inventory_result[0]['inventory_value'] if inventory_result else 0
            
            summary = {
                'total_sales': total_sales,
                'total_purchases': total_purchases,
                'gross_profit': gross_profit,
                'profit_margin': (gross_profit / total_sales * 100) if total_sales > 0 else 0,
                'inventory_value': inventory_value or 0,
                'sales_count': sales_data.get('sales_count', 0) or 0,
                'purchases_count': purchases_data.get('purchases_count', 0) or 0,
                'total_discounts': sales_data.get('total_discounts', 0) or 0,
                'total_tax': sales_data.get('total_tax', 0) or 0
            }
            
            # بيانات مفصلة للعرض
            detailed_data = [
                {'metric': 'إجمالي المبيعات', 'value': total_sales, 'type': 'income'},
                {'metric': 'إجمالي المشتريات', 'value': total_purchases, 'type': 'expense'},
                {'metric': 'الربح الإجمالي', 'value': gross_profit, 'type': 'profit'},
                {'metric': 'قيمة المخزون', 'value': inventory_value or 0, 'type': 'asset'},
                {'metric': 'إجمالي الخصومات', 'value': summary['total_discounts'], 'type': 'discount'},
                {'metric': 'إجمالي الضرائب', 'value': summary['total_tax'], 'type': 'tax'}
            ]
            
            charts_data = {
                'profit_breakdown': {
                    'المبيعات': total_sales,
                    'المشتريات': total_purchases,
                    'الربح': gross_profit
                },
                'monthly_trend': self._get_monthly_financial_trend(filters)
            }
            
            return ReportData(
                title="تقرير الملخص المالي",
                subtitle=f"من {filters.start_date.strftime('%Y-%m-%d') if filters.start_date else 'البداية'} إلى {filters.end_date.strftime('%Y-%m-%d') if filters.end_date else 'النهاية'}",
                generated_at=datetime.now(),
                filters=filters,
                data=detailed_data,
                summary=summary,
                charts_data=charts_data
            )
            
        except Exception as e:
            self.logger.error(f"خطأ في إنشاء تقرير الملخص المالي: {e}")
            raise
    
    def export_report(self, report_data: ReportData, format_type: ExportFormat, filename: Optional[str] = None) -> str:
        """تصدير التقرير بالتنسيق المحدد"""
        try:
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"report_{timestamp}.{format_type.value}"
            
            filepath = os.path.join(self.reports_dir, filename)
            
            if format_type == ExportFormat.JSON:
                return self._export_to_json(report_data, filepath)
            elif format_type == ExportFormat.CSV:
                return self._export_to_csv(report_data, filepath)
            elif format_type == ExportFormat.EXCEL and EXCEL_AVAILABLE:
                return self._export_to_excel(report_data, filepath)
            elif format_type == ExportFormat.PDF and PDF_AVAILABLE:
                return self._export_to_pdf(report_data, filepath)
            else:
                raise ValueError(f"تنسيق التصدير غير مدعوم أو المكتبات المطلوبة غير متوفرة: {format_type.value}")
            
        except Exception as e:
            self.logger.error(f"خطأ في تصدير التقرير: {e}")
            raise
    
    def _export_to_json(self, report_data: ReportData, filepath: str) -> str:
        """تصدير إلى JSON"""
        data = asdict(report_data)
        # تحويل التواريخ إلى نصوص
        data['generated_at'] = report_data.generated_at.isoformat()
        if data['filters']['start_date']:
            data['filters']['start_date'] = report_data.filters.start_date.isoformat()
        if data['filters']['end_date']:
            data['filters']['end_date'] = report_data.filters.end_date.isoformat()
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        return filepath
    
    def _export_to_csv(self, report_data: ReportData, filepath: str) -> str:
        """تصدير إلى CSV"""
        if not report_data.data:
            raise ValueError("لا توجد بيانات للتصدير")
        
        df = pd.DataFrame(report_data.data)
        df.to_csv(filepath, index=False, encoding='utf-8-sig')
        
        return filepath
    
    def _export_to_excel(self, report_data: ReportData, filepath: str) -> str:
        """تصدير إلى Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "التقرير"
        
        # إعداد الأنماط
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # كتابة العنوان
        ws['A1'] = report_data.title
        ws['A1'].font = Font(bold=True, size=16)
        ws['A2'] = report_data.subtitle
        ws['A3'] = f"تاريخ الإنشاء: {report_data.generated_at.strftime('%Y-%m-%d %H:%M:%S')}"
        
        # كتابة البيانات
        if report_data.data:
            df = pd.DataFrame(report_data.data)
            
            # إضافة البيانات بدءاً من الصف 5
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # تنسيق الرأس
            for cell in ws[5]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
        
        # إضافة ورقة الملخص
        summary_ws = wb.create_sheet("الملخص")
        row = 1
        for key, value in report_data.summary.items():
            if isinstance(value, dict):
                summary_ws[f'A{row}'] = str(key)
                summary_ws[f'A{row}'].font = Font(bold=True)
                row += 1
                for sub_key, sub_value in value.items():
                    summary_ws[f'A{row}'] = f"  {sub_key}"
                    summary_ws[f'B{row}'] = str(sub_value)
                    row += 1
            else:
                summary_ws[f'A{row}'] = str(key)
                summary_ws[f'B{row}'] = str(value)
                row += 1
        
        wb.save(filepath)
        return filepath
    
    def _export_to_pdf(self, report_data: ReportData, filepath: str) -> str:
        """تصدير إلى PDF"""
        # قالب HTML للتقرير
        html_template = """
        <!DOCTYPE html>
        <html dir="rtl" lang="ar">
        <head>
            <meta charset="UTF-8">
            <title>{{ title }}</title>
            <style>
                body { font-family: 'Arial', sans-serif; direction: rtl; }
                .header { text-align: center; margin-bottom: 30px; }
                .title { font-size: 24px; font-weight: bold; color: #333; }
                .subtitle { font-size: 16px; color: #666; margin-top: 10px; }
                .summary { background: #f5f5f5; padding: 15px; margin: 20px 0; border-radius: 5px; }
                .data-table { width: 100%; border-collapse: collapse; margin-top: 20px; }
                .data-table th, .data-table td { border: 1px solid #ddd; padding: 8px; text-align: right; }
                .data-table th { background-color: #366092; color: white; }
                .footer { margin-top: 30px; text-align: center; font-size: 12px; color: #666; }
            </style>
        </head>
        <body>
            <div class="header">
                <div class="title">{{ title }}</div>
                <div class="subtitle">{{ subtitle }}</div>
                <div>تاريخ الإنشاء: {{ generated_at }}</div>
            </div>
            
            <div class="summary">
                <h3>ملخص التقرير</h3>
                {% for key, value in summary.items() %}
                    <p><strong>{{ key }}:</strong> {{ value }}</p>
                {% endfor %}
            </div>
            
            {% if data %}
            <table class="data-table">
                <thead>
                    <tr>
                        {% for key in data[0].keys() %}
                            <th>{{ key }}</th>
                        {% endfor %}
                    </tr>
                </thead>
                <tbody>
                    {% for row in data %}
                    <tr>
                        {% for value in row.values() %}
                            <td>{{ value }}</td>
                        {% endfor %}
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
            {% endif %}
            
            <div class="footer">
                <p>تم إنشاء هذا التقرير بواسطة نظام الإصدار المنطقي</p>
            </div>
        </body>
        </html>
        """
        
        template = Template(html_template)
        html_content = template.render(
            title=report_data.title,
            subtitle=report_data.subtitle,
            generated_at=report_data.generated_at.strftime('%Y-%m-%d %H:%M:%S'),
            summary=report_data.summary,
            data=report_data.data
        )
        
        # تحويل إلى PDF
        HTML(string=html_content).write_pdf(filepath)
        return filepath
    
    def _get_daily_sales_chart_data(self, filters: ReportFilter) -> Dict[str, float]:
        """الحصول على بيانات المبيعات اليومية للرسم البياني"""
        query = """
            SELECT 
                DATE(sale_date) as sale_day,
                SUM(final_amount) as daily_total
            FROM sales 
            WHERE status != 'cancelled'
        """
        
        params = []
        if filters.start_date:
            query += " AND sale_date >= ?"
            params.append(filters.start_date.strftime('%Y-%m-%d'))
        
        if filters.end_date:
            query += " AND sale_date <= ?"
            params.append(filters.end_date.strftime('%Y-%m-%d'))
        
        query += " GROUP BY DATE(sale_date) ORDER BY sale_day"
        
        results = self.db.execute_query(query, params)
        return {row['sale_day']: row['daily_total'] for row in results}
    
    def _get_top_customers_chart_data(self, filters: ReportFilter) -> Dict[str, float]:
        """الحصول على بيانات أفضل العملاء للرسم البياني"""
        query = """
            SELECT 
                c.name as customer_name,
                SUM(s.final_amount) as total_purchases
            FROM sales s
            JOIN customers c ON s.customer_id = c.id
            WHERE s.status != 'cancelled'
        """
        
        params = []
        if filters.start_date:
            query += " AND s.sale_date >= ?"
            params.append(filters.start_date.strftime('%Y-%m-%d'))
        
        if filters.end_date:
            query += " AND s.sale_date <= ?"
            params.append(filters.end_date.strftime('%Y-%m-%d'))
        
        query += " GROUP BY c.id, c.name ORDER BY total_purchases DESC LIMIT 10"
        
        results = self.db.execute_query(query, params)
        return {row['customer_name']: row['total_purchases'] for row in results}
    
    def _get_monthly_financial_trend(self, filters: ReportFilter) -> Dict[str, Dict[str, float]]:
        """الحصول على الاتجاه المالي الشهري"""
        # المبيعات الشهرية
        sales_query = """
            SELECT 
                strftime('%Y-%m', sale_date) as month,
                SUM(final_amount) as monthly_sales
            FROM sales 
            WHERE status != 'cancelled'
        """
        
        # المشتريات الشهرية
        purchases_query = """
            SELECT 
                strftime('%Y-%m', purchase_date) as month,
                SUM(total_amount) as monthly_purchases
            FROM purchases 
            WHERE status != 'cancelled'
        """
        
        params = []
        if filters.start_date:
            sales_query += " AND sale_date >= ?"
            purchases_query += " AND purchase_date >= ?"
            params.extend([filters.start_date.strftime('%Y-%m-%d')] * 2)
        
        if filters.end_date:
            sales_query += " AND sale_date <= ?"
            purchases_query += " AND purchase_date <= ?"
            params.extend([filters.end_date.strftime('%Y-%m-%d')] * 2)
        
        sales_query += " GROUP BY strftime('%Y-%m', sale_date) ORDER BY month"
        purchases_query += " GROUP BY strftime('%Y-%m', purchase_date) ORDER BY month"
        
        sales_results = self.db.execute_query(sales_query, params[:len(params)//2] if params else [])
        purchases_results = self.db.execute_query(purchases_query, params[len(params)//2:] if params else [])
        
        # دمج النتائج
        monthly_data = {}
        
        for row in sales_results:
            month = row['month']
            if month not in monthly_data:
                monthly_data[month] = {'sales': 0, 'purchases': 0, 'profit': 0}
            monthly_data[month]['sales'] = row['monthly_sales']
        
        for row in purchases_results:
            month = row['month']
            if month not in monthly_data:
                monthly_data[month] = {'sales': 0, 'purchases': 0, 'profit': 0}
            monthly_data[month]['purchases'] = row['monthly_purchases']
        
        # حساب الأرباح
        for month_data in monthly_data.values():
            month_data['profit'] = month_data['sales'] - month_data['purchases']
        
        return monthly_data
    
    def get_available_reports(self) -> List[Dict[str, str]]:
        """الحصول على قائمة التقارير المتاحة"""
        return [
            {
                'type': ReportType.SALES_SUMMARY.value,
                'name': 'تقرير ملخص المبيعات',
                'description': 'ملخص شامل لجميع المبيعات في فترة محددة'
            },
            {
                'type': ReportType.INVENTORY_STATUS.value,
                'name': 'تقرير حالة المخزون',
                'description': 'حالة المخزون الحالية لجميع المنتجات'
            },
            {
                'type': ReportType.FINANCIAL_SUMMARY.value,
                'name': 'تقرير الملخص المالي',
                'description': 'ملخص الوضع المالي للشركة'
            },
            # تقارير المدفوعات والحسابات
            {
                'type': ReportType.PAYMENT_SUMMARY.value,
                'name': 'تقرير ملخص المدفوعات',
                'description': 'ملخص شامل لجميع المدفوعات والمقبوضات'
            },
            {
                'type': ReportType.RECEIVABLES_AGING.value,
                'name': 'تقرير أعمار الذمم المدينة',
                'description': 'تحليل أعمار المستحقات من العملاء'
            },
            {
                'type': ReportType.PAYABLES_AGING.value,
                'name': 'تقرير أعمار الذمم الدائنة',
                'description': 'تحليل أعمار المستحقات للموردين'
            },
            {
                'type': ReportType.CASH_FLOW.value,
                'name': 'تقرير التدفق النقدي',
                'description': 'تحليل التدفقات النقدية الداخلة والخارجة'
            },
            {
                'type': ReportType.PAYMENT_ANALYSIS.value,
                'name': 'تقرير تحليل المدفوعات',
                'description': 'تحليل مفصل لأنماط وسلوكيات المدفوعات'
            },
            {
                'type': ReportType.PAYMENT_METHODS_ANALYSIS.value,
                'name': 'تقرير تحليل طرق الدفع',
                'description': 'تحليل استخدام طرق الدفع المختلفة'
            }
        ]
    
    def get_supported_formats(self) -> List[Dict[str, Any]]:
        """الحصول على تنسيقات التصدير المدعومة"""
        formats = [
            {'format': ExportFormat.JSON.value, 'name': 'JSON', 'available': True},
            {'format': ExportFormat.CSV.value, 'name': 'CSV', 'available': True},
            {'format': ExportFormat.EXCEL.value, 'name': 'Excel', 'available': EXCEL_AVAILABLE},
            {'format': ExportFormat.PDF.value, 'name': 'PDF', 'available': PDF_AVAILABLE}
        ]
        
        return formats

    # ==================== تقارير المدفوعات والحسابات ====================
    
    def generate_payment_summary_report(self, filters: ReportFilter) -> ReportData:
        """توليد تقرير ملخص المدفوعات"""
        try:
            # بناء الاستعلام الأساسي
            query = """
                SELECT 
                    p.payment_date,
                    p.payment_type,
                    p.payment_method,
                    p.amount,
                    p.payment_status,
                    CASE 
                        WHEN p.payment_type = 'دفعة عميل' THEN c.name
                        WHEN p.payment_type = 'دفعة مورد' THEN s.name
                        ELSE 'غير محدد'
                    END as entity_name
                FROM payments p
                LEFT JOIN customers c ON p.entity_id = c.id AND p.payment_type = 'دفعة عميل'
                LEFT JOIN suppliers s ON p.entity_id = s.id AND p.payment_type = 'دفعة مورد'
                WHERE 1=1
            """
            
            params = []
            
            # تطبيق الفلاتر
            if filters.start_date:
                query += " AND p.payment_date >= ?"
                params.append(filters.start_date)
            
            if filters.end_date:
                query += " AND p.payment_date <= ?"
                params.append(filters.end_date)
                
            if filters.payment_type:
                query += " AND p.payment_type = ?"
                params.append(filters.payment_type)
                
            if filters.payment_method:
                query += " AND p.payment_method = ?"
                params.append(filters.payment_method)
                
            if filters.payment_status:
                query += " AND p.payment_status = ?"
                params.append(filters.payment_status)
                
            if filters.entity_id:
                query += " AND p.entity_id = ?"
                params.append(filters.entity_id)
                
            if filters.min_amount:
                query += " AND p.amount >= ?"
                params.append(filters.min_amount)
                
            if filters.max_amount:
                query += " AND p.amount <= ?"
                params.append(filters.max_amount)
            
            query += " ORDER BY p.payment_date DESC"
            
            # تنفيذ الاستعلام
            cursor = self.db_manager.connection.cursor()
            cursor.execute(query, params)
            results = cursor.fetchall()
            
            # تحويل النتائج إلى قائمة من القواميس
            data = []
            total_amount = 0
            customer_payments = 0
            supplier_payments = 0
            
            for row in results:
                payment_data = {
                    'payment_date': row[0],
                    'payment_type': row[1],
                    'payment_method': row[2],
                    'amount': float(row[3]),
                    'payment_status': row[4],
                    'entity_name': row[5]
                }
                data.append(payment_data)
                
                amount = float(row[3])
                total_amount += amount
                
                if row[1] == 'دفعة عميل':
                    customer_payments += amount
                elif row[1] == 'دفعة مورد':
                    supplier_payments += amount
            
            # إعداد الملخص
            summary = {
                'total_payments': len(data),
                'total_amount': total_amount,
                'customer_payments': customer_payments,
                'supplier_payments': supplier_payments,
                'net_cash_flow': customer_payments - supplier_payments
            }
            
            return ReportData(
                title="تقرير ملخص المدفوعات",
                subtitle=f"من {filters.start_date or 'البداية'} إلى {filters.end_date or 'النهاية'}",
                generated_at=datetime.now(),
                filters=filters,
                data=data,
                summary=summary
            )
            
        except Exception as e:
            if self.logger:
                self.logger.error(f"خطأ في توليد تقرير ملخص المدفوعات: {str(e)}")
            raise
    
    def generate_receivables_aging_report(self, filters: ReportFilter) -> ReportData:
        """توليد تقرير أعمار الذمم المدينة"""
        try:
            aging_periods = filters.aging_periods or [30, 60, 90]
            
            query = """
                SELECT 
                    c.id,
                    c.name,
                    c.phone,
                    c.email,
                    COALESCE(SUM(
                        CASE WHEN p.payment_type = 'دفعة عميل' THEN p.amount ELSE 0 END
                    ), 0) as total_payments,
                    COALESCE(SUM(
                        CASE WHEN s.customer_id = c.id THEN s.total_amount ELSE 0 END
                    ), 0) as total_sales,
                    (COALESCE(SUM(
                        CASE WHEN s.customer_id = c.id THEN s.total_amount ELSE 0 END
                    ), 0) - COALESCE(SUM(
                        CASE WHEN p.payment_type = 'دفعة عميل' THEN p.amount ELSE 0 END
                    ), 0)) as balance
                FROM customers c
                LEFT JOIN payments p ON c.id = p.entity_id
                LEFT JOIN sales s ON c.id = s.customer_id
                GROUP BY c.id, c.name, c.phone, c.email
            """
            
            if not filters.include_zero_balances:
                query += " HAVING balance > 0"
            
            query += " ORDER BY balance DESC"
            
            cursor = self.db_manager.connection.cursor()
            cursor.execute(query)
            results = cursor.fetchall()
            
            data = []
            total_balance = 0
            
            for row in results:
                customer_data = {
                    'customer_id': row[0],
                    'customer_name': row[1],
                    'phone': row[2] or '',
                    'email': row[3] or '',
                    'total_payments': float(row[4]),
                    'total_sales': float(row[5]),
                    'balance': float(row[6]),
                    'aging_0_30': 0,
                    'aging_31_60': 0,
                    'aging_61_90': 0,
                    'aging_over_90': 0
                }
                
                # حساب الأعمار (يمكن تطويرها لاحقاً بناءً على تواريخ الفواتير)
                balance = float(row[6])
                if balance > 0:
                    customer_data['aging_0_30'] = balance  # مؤقتاً
                
                data.append(customer_data)
                total_balance += balance
            
            summary = {
                'total_customers': len(data),
                'total_balance': total_balance,
                'aging_periods': aging_periods
            }
            
            return ReportData(
                title="تقرير أعمار الذمم المدينة",
                subtitle=f"كما في {datetime.now().strftime('%Y-%m-%d')}",
                generated_at=datetime.now(),
                filters=filters,
                data=data,
                summary=summary
            )
            
        except Exception as e:
            if self.logger:
                self.logger.error(f"خطأ في توليد تقرير أعمار الذمم المدينة: {str(e)}")
            raise
    
    def generate_payables_aging_report(self, filters: ReportFilter) -> ReportData:
        """توليد تقرير أعمار الذمم الدائنة"""
        try:
            aging_periods = filters.aging_periods or [30, 60, 90]
            
            query = """
                SELECT 
                    s.id,
                    s.name,
                    s.phone,
                    s.email,
                    COALESCE(SUM(
                        CASE WHEN p.payment_type = 'دفعة مورد' THEN p.amount ELSE 0 END
                    ), 0) as total_payments,
                    COALESCE(SUM(
                        CASE WHEN pur.supplier_id = s.id THEN pur.total_amount ELSE 0 END
                    ), 0) as total_purchases,
                    (COALESCE(SUM(
                        CASE WHEN pur.supplier_id = s.id THEN pur.total_amount ELSE 0 END
                    ), 0) - COALESCE(SUM(
                        CASE WHEN p.payment_type = 'دفعة مورد' THEN p.amount ELSE 0 END
                    ), 0)) as balance
                FROM suppliers s
                LEFT JOIN payments p ON s.id = p.entity_id
                LEFT JOIN purchases pur ON s.id = pur.supplier_id
                GROUP BY s.id, s.name, s.phone, s.email
            """
            
            if not filters.include_zero_balances:
                query += " HAVING balance > 0"
            
            query += " ORDER BY balance DESC"
            
            cursor = self.db_manager.connection.cursor()
            cursor.execute(query)
            results = cursor.fetchall()
            
            data = []
            total_balance = 0
            
            for row in results:
                supplier_data = {
                    'supplier_id': row[0],
                    'supplier_name': row[1],
                    'phone': row[2] or '',
                    'email': row[3] or '',
                    'total_payments': float(row[4]),
                    'total_purchases': float(row[5]),
                    'balance': float(row[6]),
                    'aging_0_30': 0,
                    'aging_31_60': 0,
                    'aging_61_90': 0,
                    'aging_over_90': 0
                }
                
                # حساب الأعمار (يمكن تطويرها لاحقاً بناءً على تواريخ الفواتير)
                balance = float(row[6])
                if balance > 0:
                    supplier_data['aging_0_30'] = balance  # مؤقتاً
                
                data.append(supplier_data)
                total_balance += balance
            
            summary = {
                'total_suppliers': len(data),
                'total_balance': total_balance,
                'aging_periods': aging_periods
            }
            
            return ReportData(
                title="تقرير أعمار الذمم الدائنة",
                subtitle=f"كما في {datetime.now().strftime('%Y-%m-%d')}",
                generated_at=datetime.now(),
                filters=filters,
                data=data,
                summary=summary
            )
            
        except Exception as e:
            if self.logger:
                self.logger.error(f"خطأ في توليد تقرير أعمار الذمم الدائنة: {str(e)}")
            raise
    
    def generate_cash_flow_report(self, filters: ReportFilter) -> ReportData:
        """توليد تقرير التدفق النقدي"""
        try:
            query = """
                SELECT 
                    DATE(p.payment_date) as payment_date,
                    p.payment_type,
                    p.payment_method,
                    SUM(p.amount) as daily_amount
                FROM payments p
                WHERE p.payment_status = 'مكتمل'
            """
            
            params = []
            
            if filters.start_date:
                query += " AND p.payment_date >= ?"
                params.append(filters.start_date)
            
            if filters.end_date:
                query += " AND p.payment_date <= ?"
                params.append(filters.end_date)
                
            if filters.payment_method:
                query += " AND p.payment_method = ?"
                params.append(filters.payment_method)
            
            query += " GROUP BY DATE(p.payment_date), p.payment_type, p.payment_method"
            query += " ORDER BY payment_date DESC"
            
            cursor = self.db_manager.connection.cursor()
            cursor.execute(query, params)
            results = cursor.fetchall()
            
            # تجميع البيانات حسب التاريخ
            daily_flow = {}
            total_inflow = 0
            total_outflow = 0
            
            for row in results:
                date = row[0]
                payment_type = row[1]
                payment_method = row[2]
                amount = float(row[3])
                
                if date not in daily_flow:
                    daily_flow[date] = {
                        'date': date,
                        'inflow': 0,
                        'outflow': 0,
                        'net': 0,
                        'methods': {}
                    }
                
                if payment_method not in daily_flow[date]['methods']:
                    daily_flow[date]['methods'][payment_method] = {
                        'inflow': 0,
                        'outflow': 0
                    }
                
                if payment_type == 'دفعة عميل':
                    daily_flow[date]['inflow'] += amount
                    daily_flow[date]['methods'][payment_method]['inflow'] += amount
                    total_inflow += amount
                elif payment_type == 'دفعة مورد':
                    daily_flow[date]['outflow'] += amount
                    daily_flow[date]['methods'][payment_method]['outflow'] += amount
                    total_outflow += amount
                
                daily_flow[date]['net'] = daily_flow[date]['inflow'] - daily_flow[date]['outflow']
            
            # تحويل إلى قائمة
            data = list(daily_flow.values())
            
            summary = {
                'total_inflow': total_inflow,
                'total_outflow': total_outflow,
                'net_cash_flow': total_inflow - total_outflow,
                'period_days': len(data)
            }
            
            return ReportData(
                title="تقرير التدفق النقدي",
                subtitle=f"من {filters.start_date or 'البداية'} إلى {filters.end_date or 'النهاية'}",
                generated_at=datetime.now(),
                filters=filters,
                data=data,
                summary=summary
            )
            
        except Exception as e:
            if self.logger:
                self.logger.error(f"خطأ في توليد تقرير التدفق النقدي: {str(e)}")
            raise
    
    def generate_payment_analysis_report(self, filters: ReportFilter) -> ReportData:
        """توليد تقرير تحليل المدفوعات"""
        try:
            # تحليل المدفوعات حسب النوع والطريقة
            query = """
                SELECT 
                    p.payment_type,
                    p.payment_method,
                    COUNT(*) as transaction_count,
                    SUM(p.amount) as total_amount,
                    AVG(p.amount) as average_amount,
                    MIN(p.amount) as min_amount,
                    MAX(p.amount) as max_amount
                FROM payments p
                WHERE p.payment_status = 'مكتمل'
            """
            
            params = []
            
            if filters.start_date:
                query += " AND p.payment_date >= ?"
                params.append(filters.start_date)
            
            if filters.end_date:
                query += " AND p.payment_date <= ?"
                params.append(filters.end_date)
            
            query += " GROUP BY p.payment_type, p.payment_method"
            query += " ORDER BY total_amount DESC"
            
            cursor = self.db_manager.connection.cursor()
            cursor.execute(query, params)
            results = cursor.fetchall()
            
            data = []
            total_transactions = 0
            total_amount = 0
            
            for row in results:
                analysis_data = {
                    'payment_type': row[0],
                    'payment_method': row[1],
                    'transaction_count': row[2],
                    'total_amount': float(row[3]),
                    'average_amount': float(row[4]),
                    'min_amount': float(row[5]),
                    'max_amount': float(row[6]),
                    'percentage': 0  # سيتم حسابها لاحقاً
                }
                data.append(analysis_data)
                total_transactions += row[2]
                total_amount += float(row[3])
            
            # حساب النسب المئوية
            for item in data:
                if total_amount > 0:
                    item['percentage'] = (item['total_amount'] / total_amount) * 100
            
            summary = {
                'total_transactions': total_transactions,
                'total_amount': total_amount,
                'average_transaction': total_amount / total_transactions if total_transactions > 0 else 0
            }
            
            return ReportData(
                title="تقرير تحليل المدفوعات",
                subtitle=f"من {filters.start_date or 'البداية'} إلى {filters.end_date or 'النهاية'}",
                generated_at=datetime.now(),
                filters=filters,
                data=data,
                summary=summary
            )
            
        except Exception as e:
            if self.logger:
                self.logger.error(f"خطأ في توليد تقرير تحليل المدفوعات: {str(e)}")
            raise
    
    def generate_payment_methods_analysis_report(self, filters: ReportFilter) -> ReportData:
        """توليد تقرير تحليل طرق الدفع"""
        try:
            query = """
                SELECT 
                    p.payment_method,
                    COUNT(*) as usage_count,
                    SUM(p.amount) as total_amount,
                    AVG(p.amount) as average_amount,
                    COUNT(CASE WHEN p.payment_type = 'دفعة عميل' THEN 1 END) as customer_payments,
                    COUNT(CASE WHEN p.payment_type = 'دفعة مورد' THEN 1 END) as supplier_payments
                FROM payments p
                WHERE p.payment_status = 'مكتمل'
            """
            
            params = []
            
            if filters.start_date:
                query += " AND p.payment_date >= ?"
                params.append(filters.start_date)
            
            if filters.end_date:
                query += " AND p.payment_date <= ?"
                params.append(filters.end_date)
            
            query += " GROUP BY p.payment_method"
            query += " ORDER BY total_amount DESC"
            
            cursor = self.db_manager.connection.cursor()
            cursor.execute(query, params)
            results = cursor.fetchall()
            
            data = []
            total_usage = 0
            total_amount = 0
            
            for row in results:
                method_data = {
                    'payment_method': row[0],
                    'usage_count': row[1],
                    'total_amount': float(row[2]),
                    'average_amount': float(row[3]),
                    'customer_payments': row[4],
                    'supplier_payments': row[5],
                    'usage_percentage': 0,
                    'amount_percentage': 0
                }
                data.append(method_data)
                total_usage += row[1]
                total_amount += float(row[2])
            
            # حساب النسب المئوية
            for item in data:
                if total_usage > 0:
                    item['usage_percentage'] = (item['usage_count'] / total_usage) * 100
                if total_amount > 0:
                    item['amount_percentage'] = (item['total_amount'] / total_amount) * 100
            
            summary = {
                'total_methods': len(data),
                'total_usage': total_usage,
                'total_amount': total_amount,
                'most_used_method': data[0]['payment_method'] if data else None,
                'highest_amount_method': max(data, key=lambda x: x['total_amount'])['payment_method'] if data else None
            }
            
            return ReportData(
                title="تقرير تحليل طرق الدفع",
                subtitle=f"من {filters.start_date or 'البداية'} إلى {filters.end_date or 'النهاية'}",
                generated_at=datetime.now(),
                filters=filters,
                data=data,
                summary=summary
            )
            
        except Exception as e:
            if self.logger:
                self.logger.error(f"خطأ في توليد تقرير تحليل طرق الدفع: {str(e)}")
            raise